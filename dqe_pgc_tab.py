"""
DQE PGC Tab Module

Handles the DQE PGC tab interface and functionality for the QGIS plugin.
Extracted from the main dialog file for better modularity.
"""

import json
import time
import uuid
from typing import List, Dict, Any

from .compat import (
    QTimer, QVBoxLayout, QHBoxLayout, QFormLayout, QGroupBox,
    QPushButton, QApplication, QMessageBox,
    MSGBOX_ACCEPTROLE, MSGBOX_REJECTROLE, QGIS_SUCCESS
)
from .base_tab import BaseDQETab
from qgis.core import QgsProject
from qgis.utils import iface
from .ui_components import SROComboBox, TronconComboBox, ProgressWidget
from .layer_manager import LayerManager
from .database_operations import DatabaseOperations
from .excel_manager import ExcelManager
from .workers import DQEWorker
from .dqe_utils import _db_manager, _logger, _validator, QtCompatibility
from .telemetry import send_telemetry


class DQEPGCTab(BaseDQETab):
    """Interface et logique pour l'onglet DQE PGC"""
    TAB_LABEL = "PGC"
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.last_dqe_results = None
        self.last_sro = None
        self.last_troncon = None
        self.redevance_mode_gestionnaire = False
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(4)
        config_group = QGroupBox("Configuration DQE PGC")
        config_layout = QFormLayout(config_group)
        
        self.sro_input = SROComboBox()
        self.sro_input.setToolTip("Saisir le code SRO pour charger les tronçons disponibles\nFormat attendu : XXX/XXX/XXX/XXX")
        config_layout.addRow("SRO:", self.sro_input)
        
        self.troncon_combo = TronconComboBox()
        self.troncon_combo.setEnabled(False)
        self.troncon_combo.setToolTip("Sélectionner le tronçon à traiter parmi ceux disponibles pour ce SRO")
        config_layout.addRow("Tronçon:", self.troncon_combo)
        
        layout.addWidget(config_group)
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(8)  # Espacement réduit entre boutons
        
        self.execute_button = QPushButton("Exécuter DQE PGC")
        self.execute_button.setEnabled(False)
        self.execute_button.setToolTip("Lancer le calcul DQE PGC avec attribution gestionnaire\nCharge automatiquement les couches QGIS et génère l'Excel")
        self.execute_button.clicked.connect(self.execute_dqe_pgc)
        buttons_layout.addWidget(self.execute_button)
        
        self.regenerate_excel_button = QPushButton("Régénérer Excel")
        self.regenerate_excel_button.setToolTip("Régénérer l'Excel avec les données modifiées manuellement\nDétecte automatiquement les corrections dans la couche gestionnaire")
        self.regenerate_excel_button.clicked.connect(self.regenerate_excel)
        self.regenerate_excel_button.setEnabled(False)  # Désactivé par défaut
        buttons_layout.addWidget(self.regenerate_excel_button)
        
        self.validate_button = QPushButton("Valider DQE")
        self.validate_button.setToolTip("Sauvegarder les résultats DQE en base de données\nConfirme et finalise le traitement")
        self.validate_button.clicked.connect(self.validate_dqe_pgc)
        buttons_layout.addWidget(self.validate_button)
        
        layout.addLayout(buttons_layout)
        self.progress_widget = ProgressWidget()
        layout.addWidget(self.progress_widget)
        self.sro_input.lineEdit().textChanged.connect(self.on_sro_changed)
        self.troncon_combo.currentTextChanged.connect(self.on_troncon_changed)
        self.connect_gestionnaire_layer_signals()
    
    def on_sro_changed(self):
        self._reset_plugin_state()
        
        sro = self.sro_input.lineEdit().text().strip()
        
        self.troncon_combo.clear()
        self.troncon_combo.setEnabled(False)
        self.execute_button.setEnabled(False)
        
        if len(sro) >= 5:
            QTimer.singleShot(1000, lambda: self.validate_sro_async(sro))
    
    def on_troncon_changed(self):
        """Appelé quand le tronçon change - réinitialise l'état du plugin"""
        troncon = self.troncon_combo.currentText().strip()
        
        if troncon:  # Seulement si un tronçon est sélectionné
            print(f"CHANGEMENT: Nouveau tronçon sélectionné: {troncon}")
            self._reset_plugin_state()
    
    def _reset_plugin_state(self):
        """Réinitialise complètement l'état du plugin pour un nouveau SRO/tronçon"""
        print("RÉINITIALISATION: État du plugin réinitialisé")
        self.last_dqe_results = None
        self.last_sro = None
        self.last_troncon = None
        self.redevance_mode_gestionnaire = False
        self.regenerate_excel_button.setEnabled(False)
        self.regenerate_excel_button.setText("Régénérer Excel")
        self._clean_previous_layers()
        if hasattr(self, 'progress_widget'):
            self.progress_widget.is_running = False
            self.progress_widget.progress_bar.setVisible(False)
            self.progress_widget.cancel_button.setVisible(False)
            self.progress_widget.status_label.setText("")
        
        print("RÉINITIALISATION: État du plugin réinitialisé")
    
    def _clean_previous_layers(self):
        """Nettoie les couches des traitements précédents"""
        try:
            from qgis.core import QgsProject
            project = QgsProject.instance()
            if self.layer_group:
                project.layerTreeRoot().removeChildNode(self.layer_group)
                self.layer_group = None
            self.layers_loaded.clear()
            
            print("NETTOYAGE: Couches précédentes supprimées")
            
        except Exception as e:
            print(f"ATTENTION: Erreur nettoyage couches: {e}")
    
    def validate_sro_async(self, sro: str):
        """Validation format uniquement sur frappe (pas de DB).

        Le chargement des troncons (query DB) n'est déclenché qu'une
        fois le format validé, pas à chaque frappe.
        """
        if sro != self.sro_input.lineEdit().text().strip():
            return
        
        if _validator:
            if _validator._validate_sro_format(sro):
                self.troncon_combo.set_sro(sro)
                self.troncon_combo.setEnabled(True)
                self.execute_button.setEnabled(True)
    
    def execute_dqe_pgc(self):
        """Exécution DQE PGC avec interface uniforme"""
        sro = self.sro_input.lineEdit().text().strip()
        troncon = self.troncon_combo.currentText().strip()
        
        if not sro or not troncon:
            print("ERREUR: Veuillez remplir le SRO et sélectionner un tronçon")
            return
        
        if _validator:
            is_valid, message = _validator.validate_sro_exists(sro)
            if not is_valid:
                QMessageBox.warning(self, "SRO invalide", f"Le SRO saisi n'est pas valide:\n{message}")
                return
        
        self.execute_button.setEnabled(False)
        self._exec_start = time.monotonic()
        
        try:
            worker = DQEWorker("PGC", sro, None, troncon)
            self._setup_worker_thread(worker, self.on_dqe_pgc_finished)
            
        except Exception as e:
            error_msg = f"Erreur initialisation DQE PGC: {str(e)}"
            print(f"ERREUR DQE PGC: {error_msg}")
            self.progress_widget.complete_operation(False, error_msg)
            self.execute_button.setEnabled(True)
    
    def on_dqe_pgc_finished(self, success: bool, results, message: str):
        """Callback appelé quand le traitement DQE PGC est terminé"""
        try:
            if hasattr(self, 'progress_timer'):
                self.post_processing = True
                
            if success and results:
                sro = self.sro_input.lineEdit().text().strip()
                troncon_safe = self.troncon_combo.currentText().strip().replace('/', '_')
                self.last_dqe_results = results
                self.last_sro = sro  
                self.last_troncon = troncon_safe
                print(f"Resultats DQE PGC stockes pour regeneration (SRO: {sro}, Troncon: {troncon_safe})")
                self.smooth_progress_to(92, "Création des couches...")
                QApplication.processEvents()
                current_date = time.strftime("%Y-%m-%d_%H%M%S")
                sro = self.sro_input.lineEdit().text().strip()
                troncon_safe = self.troncon_combo.currentText().strip().replace('/', '_')
                sro_safe = sro.replace('/', '_')
                group_name = f"DQE_PGC_{sro_safe}_{troncon_safe}_{current_date}"
                self.layer_group = LayerManager.create_layer_group(group_name)
                created_layers = self._load_organized_layers(results, sro, troncon_safe)
                
                if self._is_cancelled:
                    self.progress_widget.complete_operation(False, "Opération annulée")
                    send_telemetry(
                        action="execution", mode="PGC",
                        sro=sro, troncon=troncon_safe,
                        projet_code="GC",
                        verdict="cancelled", cancelled=True,
                        elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
                    )
                    return
                
                print(f"\n=== CHOIX MODE CALCUL REDEVANCE ===")
                msgBox = QMessageBox(self)
                msgBox.setWindowTitle("Mode de Calcul des Redevances")
                main_text = (
                    "<p style='font-size: 14px; margin-bottom: 15px;'>"
                    "Sélectionnez le mode de traitement des redevances :"
                    "</p>"
                    
                    "<h3 style='color: #2e7d32; font-weight: bold; margin: 10px 0 5px 0;'>"
                    "Mode Gestionnaire (Recommandé)"
                    "</h3>"
                    "<ul style='margin: 0; padding-left: 20px; line-height: 1.6;'>"
                    "<li>Calcul précis par algorithme de proximité</li>"
                    "<li>Corrections manuelles possibles</li>"
                    "<li>Couche éditable dans QGIS</li>"
                    "<li>Résultats optimisés et vérifiables</li>"
                    "</ul>"
                    
                    "<h3 style='color: #455a64; font-weight: bold; margin: 15px 0 5px 0;'>"
                    "Mode Direct (Rapide)"
                    "</h3>"
                    "<ul style='margin: 0; padding-left: 20px; line-height: 1.6;'>"
                    "<li>Traitement sans intervention</li>"
                    "<li>Données existantes uniquement</li>"
                    "<li>Export Excel automatique</li>"
                    "<li>Pas de correction possible</li>"
                    "</ul>"
                )
                msgBox.setText(main_text)
                QtCompatibility.set_rich_text_format(msgBox)
                gestionnaire_btn = msgBox.addButton("Mode Gestionnaire", MSGBOX_ACCEPTROLE)
                direct_btn = msgBox.addButton("Mode Direct", MSGBOX_REJECTROLE)
                msgBox.setDefaultButton(gestionnaire_btn)
                
                from .compat import exec_dialog
                result = exec_dialog(msgBox)
                clicked_button = msgBox.clickedButton()
                if clicked_button == gestionnaire_btn:
                    self.redevance_mode_gestionnaire = True
                elif clicked_button == direct_btn:
                    self.redevance_mode_gestionnaire = False
                else:
                    accepted_value = QtCompatibility.get_message_box_accepted()
                    self.redevance_mode_gestionnaire = (result == accepted_value)
                
                if self.redevance_mode_gestionnaire:
                    print(f" Mode GESTIONNAIRE sélectionné → Calcul précis avec corrections possibles")
                    print(f"\n=== CHARGEMENT COUCHE GESTIONNAIRE ===")
                    gestionnaire_group = LayerManager.create_layer_subgroup(self.layer_group, "Gestionnaires")
                    
                    gestionnaire_layer = LayerManager.load_gestionnaire_layer(sro, troncon_safe, gestionnaire_group)
                    if gestionnaire_layer:
                        created_layers.append(gestionnaire_layer)
                        self.layers_loaded.append(gestionnaire_layer)
                        print(f" Couche gestionnaire chargée: {gestionnaire_layer.featureCount()} segments")
                        print(f"   → Placée dans sous-groupe 'Gestionnaires'")
                        print(f"   → Permet corrections manuelles des attributions cm_gest_do")
                    else:
                        print(f"Impossible de charger la couche gestionnaire")
                else:
                    print(f" Mode DIRECT sélectionné → Utilisation données existantes")
                    print(f"   → Aucun sous-groupe 'Gestionnaires' créé")
                self.update_buttons_state()
                if hasattr(self, 'redevance_mode_gestionnaire') and self.redevance_mode_gestionnaire:
                    self.connect_gestionnaire_layer_signals()
                self.last_dqe_results = results
                print(f"\n=== GÉNÉRATION EXCEL PGC ===")
                
                self.smooth_progress_to(98, "Génération du rapport Excel...")
                QApplication.processEvents()
                excel_path = ExcelManager.create_excel_report(results, sro, "PGC", troncon_safe)
                
                if excel_path:
                    print(f" Rapport Excel généré: {excel_path}")
                    try:
                        import os
                        os.startfile(excel_path)
                        print(" Fichier Excel ouvert automatiquement")
                    except Exception as e:
                        print(f"Impossible d'ouvrir automatiquement le fichier: {e}")
                else:
                    print("Erreur lors de la génération du rapport Excel")
                self.smooth_progress_to(100, "Finalisation...")
                QApplication.processEvents()
                
                # Collapse tous les groupes pour une vue compacte
                if self.layer_group:
                    LayerManager.collapse_group_recursive(self.layer_group)
                
                final_message = f"DQE PGC terminé: {len(created_layers)} couches créées"
                self.progress_widget.complete_operation(True, final_message)
                redevance_mode_str = "gestionnaire" if getattr(self, 'redevance_mode_gestionnaire', False) else "direct"
                send_telemetry(
                    action="execution", mode="PGC",
                    sro=sro, troncon=troncon_safe,
                    projet_code="GC",
                    verdict="success",
                    n_results=len(results) if results else 0,
                    n_layers=len(created_layers) if created_layers else 0,
                    redevance_mode=redevance_mode_str,
                    excel_generated=bool(excel_path),
                    elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
                )
                
                if iface:
                    iface.messageBar().pushMessage(
                        "DQE PGC", final_message,
                        level=QGIS_SUCCESS, duration=5
                    )
            else:
                self.progress_widget.complete_operation(False, message)
                print(f"Erreur: {message}")
                send_telemetry(
                    action="execution", mode="PGC",
                    sro=self.sro_input.lineEdit().text().strip(),
                    troncon=self.troncon_combo.currentText().strip().replace('/', '_') if hasattr(self, 'troncon_combo') else "",
                    projet_code="GC",
                    verdict="failure", error_msg=message,
                    elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
                )
                
        except Exception as e:
            error_msg = f"Erreur post-traitement DQE PGC: {str(e)}"
            print(f" {error_msg}")
            import traceback
            print(traceback.format_exc())
            self.progress_widget.complete_operation(False, error_msg)
            send_telemetry(
                action="execution", mode="PGC",
                sro=self.sro_input.lineEdit().text().strip(),
                troncon=self.troncon_combo.currentText().strip().replace('/', '_') if hasattr(self, 'troncon_combo') else "",
                projet_code="GC",
                verdict="failure", error_msg=error_msg,
                elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
            )
        finally:
            self._cleanup_thread()
            self.execute_button.setEnabled(True)
    
    def validate_dqe_pgc(self):
        """Valide et sauvegarde le DQE PGC dans la base de données
        Structure: 1 ligne dqe_result + couches avec geometries
        """
        validate_start = time.monotonic()
        try:
            sro = self.sro_input.currentText().strip()
            if not sro:
                QMessageBox.warning(self, "Validation DQE", "Veuillez sélectionner un SRO")
                return
            
            if not self.last_dqe_results:
                QMessageBox.warning(self, "Validation DQE", 
                    "Aucun résultat DQE trouvé.\nVeuillez d'abord exécuter le DQE PGC.")
                return
            
            if not self.layer_group:
                QMessageBox.warning(self, "Validation DQE", "Aucune couche DQE PGC chargée à valider")
                return
            
            projet_code = "GC"  # Génie Civil PGC
            user_name = _db_manager.config.user if _db_manager and _db_manager.config else "unknown"
            troncon = self.troncon_combo.currentText().strip() if hasattr(self, 'troncon_combo') else ''
            
            # Construire le tableau JSON - TOUTES les lignes dans l'ordre exact
            dqe_data = []
            for result in self.last_dqe_results:
                if hasattr(result, 'designation'):
                    designation = result.designation or ''
                    quantite = getattr(result, 'quantite', 0)
                    unite = getattr(result, 'unite', '') or ''
                    ids = getattr(result, 'ids', '') or ''
                else:
                    designation = result.get('designation') or result.get('Désignation') or ''
                    quantite = result.get('quantite') or result.get('Quantité') or 0
                    unite = result.get('unite') or result.get('Unité') or ''
                    ids = result.get('ids') or result.get('Ids') or ''
                
                try:
                    quantite_num = float(quantite) if quantite is not None else 0
                except (ValueError, TypeError):
                    quantite_num = 0
                
                dqe_data.append({
                    'designation': designation,
                    'quantite': quantite_num,
                    'unite': unite,
                    'ids': str(ids) if ids else None
                })
            
            if not dqe_data:
                QMessageBox.warning(self, "Validation DQE", "Aucune donnée DQE à enregistrer")
                return
            
            print(f"Validation DQE PGC: {len(dqe_data)} lignes à enregistrer")
            
            # Sauvegarde dqe_result
            with _db_manager.get_cursor() as cursor:
                query = """
                    INSERT INTO dqe.dqejson 
                    (sro, nom_dqe, projet, categorie, champs, user_name, version_projet) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    sro,
                    f"DQE_PGC_{sro}_{troncon}".replace('/', '_'),
                    projet_code,
                    'dqe_result',
                    json.dumps(dqe_data),
                    user_name,
                    None  # Version auto-assignée par trigger
                ))
            
            print(f"Validation: 1 ligne dqe_result avec {len(dqe_data)} elements")
            
            # Sauvegarde de TOUTES les couches avec géométries
            layers_count = 0
            if self.layer_group:
                layers_count = LayerManager.save_layers_to_db(
                    self.layer_group, sro, f"DQE_PGC_{sro}",
                    projet_code, user_name, _db_manager
                )
            
            total_saved = len(dqe_data) + layers_count
            print(f"Validation terminee: {len(dqe_data)} resultats SQL + {layers_count} couches")
            
            QMessageBox.information(
                self, 
                "Validation DQE PGC", 
                f"Validation terminee!\n\n"
                f"- SRO: {sro}\n"
                f"- Troncon: {troncon}\n"
                f"- Type: {projet_code}\n"
                f"- Resultats DQE: {len(dqe_data)}\n"
                f"- Couches archivees: {layers_count}\n"
                f"- Total: {total_saved} elements"
            )
            
            if _logger:
                _logger.info(f"DQE PGC validé - SRO: {sro}, Type: {projet_code}, Couches: {layers_count}, DQE: {len(dqe_data)}")
            send_telemetry(
                action="validation", mode="PGC",
                sro=sro, troncon=troncon, projet_code=projet_code,
                verdict="success",
                n_dqe_data=len(dqe_data), n_layers_saved=layers_count,
                redevance_mode="gestionnaire" if getattr(self, 'redevance_mode_gestionnaire', False) else "direct",
                elapsed_ms=int((time.monotonic() - validate_start) * 1000),
            )
            
        except Exception as e:
            error_msg = f"Erreur validation DQE PGC: {str(e)}"
            print(f"{error_msg}")
            import traceback
            print(traceback.format_exc())
            if _logger:
                _logger.error(error_msg, exception=e)
            QMessageBox.critical(self, "Erreur", f"Erreur validation DQE PGC: {str(e)}")
            send_telemetry(
                action="validation", mode="PGC",
                sro=self.sro_input.currentText().strip() if hasattr(self, 'sro_input') else "",
                troncon=self.troncon_combo.currentText().strip() if hasattr(self, 'troncon_combo') else "",
                projet_code="GC",
                verdict="failure", error_msg=str(e),
                elapsed_ms=int((time.monotonic() - validate_start) * 1000),
            )
    
    def regenerate_excel(self):
        """Régénère le fichier Excel avec les données de redevance selon le mode sélectionné"""
        if not hasattr(self, 'redevance_mode_gestionnaire') or not self.redevance_mode_gestionnaire:
            print("INFORMATION: Régénération Excel non disponible en mode direct")
            return
        sro = self.sro_input.lineEdit().text().strip()
        troncon = self.troncon_combo.currentText().strip()
        
        if not sro or not troncon:
            print("ERREUR: Veuillez remplir le SRO et sélectionner un tronçon")
            return
            
        print(f"\n=== RÉGÉNÉRATION EXCEL (Mode Gestionnaire) ===")
        print(f"SRO: {sro}, Tronçon: {troncon}")
        
        try:
            modified_data = self.get_modified_gestionnaire_data(sro, troncon)
            
            if modified_data is None:
                print("Impossible de récupérer les données gestionnaire modifiées")
                return
            
            print(f" Données gestionnaire récupérées: {len(modified_data)} segments")
            from .database_operations import DatabaseOperations
            redevance_data = DatabaseOperations.get_redevance_from_modified_gestionnaire(
                sro, troncon, modified_data
            )
            
            if not redevance_data:
                print(" Aucune donnée de redevance calculée")
                return
            
            print(f" Redevances calculées avec données modifiées")
            self._generate_excel_file(sro, troncon, redevance_data, suffix="_gestionnaire_modifié")
            
            print("Excel regenere avec succes avec les modifications gestionnaire")
            
        except Exception as e:
            print(f" Erreur lors de la régénération Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(
                self, 
                "Erreur régénération Excel", 
                f"Impossible de régénérer le fichier Excel:\n\n{str(e)}\n\nVérifiez que:\n- La couche gestionnaire est bien chargée\n- Vous avez les droits d'écriture dans le dossier\n- Le fichier Excel n'est pas ouvert dans un autre programme"
            )
    
    def update_buttons_state(self):
        """Met à jour l'état des boutons selon le mode redevance et les modifications"""
        if hasattr(self, 'redevance_mode_gestionnaire') and self.redevance_mode_gestionnaire:
            self.regenerate_excel_button.setEnabled(True)
            self.regenerate_excel_button.setText("Régénérer Excel (Mode Gestionnaire)")
            print(" Bouton Excel activé - Mode Gestionnaire")
        else:
            self.regenerate_excel_button.setEnabled(False)
            self.regenerate_excel_button.setText("Régénérer Excel (Non disponible en mode direct)")
            print(" Bouton Excel désactivé - Mode Direct")
    
    def connect_gestionnaire_layer_signals(self):
        """Connecte les signaux de modification de la couche gestionnaire"""
        if not hasattr(self, 'redevance_mode_gestionnaire') or not self.redevance_mode_gestionnaire:
            return
        project = QgsProject.instance()
        gestionnaire_layer = None
        
        for layer in project.mapLayers().values():
            if hasattr(layer, 'name') and 'gestionnaire' in layer.name().lower():
                gestionnaire_layer = layer
                break
        
        if gestionnaire_layer and hasattr(gestionnaire_layer, 'editingStarted'):
            try:
                gestionnaire_layer.editingStarted.connect(self.on_gestionnaire_editing_started)
                gestionnaire_layer.featureAdded.connect(self.on_gestionnaire_modified)
                gestionnaire_layer.featureDeleted.connect(self.on_gestionnaire_modified)
                gestionnaire_layer.attributeValueChanged.connect(self.on_gestionnaire_modified)
                print("CONNEXION: Signaux de modification de la couche gestionnaire connectés")
            except Exception as e:
                print(f"ATTENTION: Erreur connexion signaux: {e}")
    
    def on_gestionnaire_editing_started(self):
        """Appelé quand l'édition de la couche gestionnaire commence"""
        print("ÉDITION: Mode édition activé pour la couche gestionnaire")
        self.regenerate_excel_button.setEnabled(True)
        self.regenerate_excel_button.setText("Régénérer Excel (Modifications détectées)")
    
    def on_gestionnaire_modified(self):
        """Appelé quand la couche gestionnaire est modifiée"""
        print("MODIFICATION: Changements détectés dans la couche gestionnaire")
        self.regenerate_excel_button.setEnabled(True)
        self.regenerate_excel_button.setText("Régénérer Excel (Modifications détectées)")
    
    def get_modified_gestionnaire_data(self, sro, troncon):
        """Récupère les données modifiées de la couche gestionnaire si elle existe et a été modifiée"""
        try:
            gestionnaire_layer = None
            layer_name_pattern = f"Gestionnaire - {sro} - {troncon}"
            
            project = QgsProject.instance()
            matching_layers = []
            for layer in project.mapLayers().values():
                if layer and layer.isValid() and layer_name_pattern in layer.name():
                    matching_layers.append(layer)
            gestionnaire_layer = None
            if matching_layers:
                def get_suffix_number(layer_name):
                    if layer_name == layer_name_pattern:
                        return 0  # Pas de suffixe = priorité maximale
                    import re
                    match = re.search(r'\((\d+)\)$', layer_name)
                    return int(match.group(1)) if match else 999
                sorted_layers = sorted(matching_layers, key=lambda l: get_suffix_number(l.name()), reverse=True)
                gestionnaire_layer = sorted_layers[0]
            
            if matching_layers:
                print(f" {len(matching_layers)} couche(s) gestionnaire trouvée(s) pour le pattern '{layer_name_pattern}'")
                for i, layer in enumerate(matching_layers):
                    marker = " ← SÉLECTIONNÉE" if layer == gestionnaire_layer else ""
                    print(f"   {i+1}. {layer.name()}{marker}")
            
            if not gestionnaire_layer:
                print(" Couche gestionnaire non trouvée - utilisation des données originales")
                return None
            
            print(f" Récupération des données modifiées de la couche gestionnaire")
            print(f"   Couche: {gestionnaire_layer.name()}")
            print(f"   Entités: {gestionnaire_layer.featureCount()}")
            modified_data = []
            for feature in gestionnaire_layer.getFeatures():
                row_data = {
                    'troncon_gid': feature.attribute('troncon_gid'),
                    'segment_id': feature.attribute('segment_id'),
                    'cm_gest_do': feature.attribute('cm_gest_do'),
                    'cm_compo': feature.attribute('cm_compo'),
                    'long': feature.attribute('long'),
                    'distance_route_m': feature.attribute('distance_route_m'),
                    'angle_parallelisme_deg': feature.attribute('angle_parallelisme_deg'),
                    'confiance_niveau': feature.attribute('confiance_niveau'),
                    'methode_attribution': feature.attribute('methode_attribution'),
                    'nb_pot_ac': feature.attribute('nb_pot_ac')
                }
                modified_data.append(row_data)
            
            print(f" {len(modified_data)} lignes récupérées de la couche gestionnaire")
            return modified_data
            
        except Exception as e:
            print(f"Erreur récupération données gestionnaire: {str(e)}")
            return None
    
    def _auto_generate_excel_direct_mode(self, sro, troncon, results):
        """Génère automatiquement l'Excel en mode direct"""
        try:
            from .database_operations import DatabaseOperations
            redevance_data = DatabaseOperations.get_redevance_from_results(sro, troncon, results)
            
            if not redevance_data:
                print(" Aucune donnée de redevance calculée")
                return
            
            print(f" Redevances calculées avec données existantes")
            self._generate_excel_file(sro, troncon, redevance_data, suffix="_direct")
            
            print("Excel genere automatiquement avec succes en mode direct")
            
        except Exception as e:
            print(f"Erreur lors de la génération automatique Excel en mode direct: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(
                self, 
                "Erreur génération Excel", 
                f"Impossible de générer le fichier Excel:\n\n{str(e)}\n\nVérifiez que:\n- Vous avez les droits d'écriture dans le dossier\n- Le fichier Excel n'est pas ouvert dans un autre programme"
            )
    
    def _generate_excel_file(self, sro, troncon, redevance_data, suffix=""):
        """Génère un fichier Excel avec le template PGC et les données de redevance"""
        try:
            import os
            import tempfile
            import pandas as pd
            from datetime import datetime
            from openpyxl import load_workbook
            import shutil
            
            print(f"Generation Excel avec template PGC: {len(redevance_data)} entrees de redevance")
            rows = []
            for entry in redevance_data:
                if isinstance(entry, dict):
                    if 'concessionnaire_voirie' in entry:
                        row = {'concessionnaire_voirie': entry['concessionnaire_voirie']}
                        total_redevance = 0
                        for key, value in entry.items():
                            if key != 'concessionnaire_voirie':
                                if hasattr(value, '__float__'):
                                    value = float(value)
                                row[key] = value
                                if not ('poteau' in key.lower() or 'nb_unites' in key.lower()):
                                    total_redevance += value if isinstance(value, (int, float)) else 0
                        row['total_redevance'] = total_redevance
                        rows.append(row)
                    elif 'redevance_data' in entry:
                        redevance_dict = entry['redevance_data']
                        row = {
                            'concessionnaire_voirie': entry.get('gest_do', ''),
                            'total_redevance': entry.get('total_redevance', 0)
                        }
                        for compo, montant in redevance_dict.items():
                            if hasattr(montant, '__float__'):
                                montant = float(montant)
                            row[compo] = montant
                        rows.append(row)
            
            if not rows:
                print(" Aucune donnée à exporter")
                return
            df_redevance = pd.DataFrame(rows)
            print(f" DataFrame redevance créé: {len(df_redevance)} lignes, colonnes: {list(df_redevance.columns)}")
            dqe_pgc_data = []
            if hasattr(self, 'last_dqe_results') and self.last_dqe_results:
                for result in self.last_dqe_results:
                    dqe_pgc_data.append({
                        'Désignation': result.designation,
                        'Unité': getattr(result, 'unite', ''),
                        'Quantité': result.quantite
                    })
            
            df_dqe_pgc = pd.DataFrame(dqe_pgc_data) if dqe_pgc_data else pd.DataFrame([{
                'Désignation': 'Aucune donnée', 'Unité': '', 'Quantité': 0
            }])
            print(f" DataFrame DQE PGC créé: {len(df_dqe_pgc)} lignes")
            plugin_dir = os.path.dirname(__file__)
            template_path = os.path.join(plugin_dir, 'files', 'template_dqe_pgc.xlsx')
            
            if not os.path.exists(template_path):
                print(f" Template PGC non trouvé: {template_path}")
                return
            
            print(f" Template PGC trouvé: {template_path}")
            def clean_filename(name):
                invalid_chars = ['\\', '/', ':', '*', '?', '\\"', '<', '>', '|']
                for char in invalid_chars:
                    name = name.replace(char, '_')
                return name
            
            sro_clean = clean_filename(sro)
            troncon_clean = clean_filename(troncon)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"redevance_{sro_clean}_{troncon_clean}{suffix}_{timestamp}.xlsx"
            temp_dir = tempfile.gettempdir()
            filepath = os.path.join(temp_dir, filename)
            shutil.copy2(template_path, filepath)
            print(f"Template copié vers: {filepath}")
            workbook = load_workbook(filepath)
            if 'DQE PGC' in workbook.sheetnames:
                dqe_sheet = workbook['DQE PGC']
                print("Feuille 'DQE PGC' trouvée dans template")
                from .excel_manager import ExcelManager
                ExcelManager._fill_pgc_template(dqe_sheet, df_dqe_pgc, sro, troncon, workbook)
            else:
                print("Feuille 'DQE PGC' non trouvée dans template")
            if 'REDEVANCE' in workbook.sheetnames:
                redevance_sheet = workbook['REDEVANCE']
                print("Feuille 'REDEVANCE' trouvée dans template")
                self._fill_template_redevance(redevance_sheet, df_redevance)
            else:
                print("Feuille 'REDEVANCE' non trouvée dans template")
                redevance_sheet = workbook.create_sheet(title='REDEVANCE')
                print("Feuille 'REDEVANCE' créée")
                self._fill_template_redevance(redevance_sheet, df_redevance)
            try:
                workbook.save(filepath)
                workbook.close()
                print(f"Fichier Excel sauvegardé avec succès: {filepath}")
            except Exception as save_error:
                print(f"Erreur lors de la sauvegarde Excel (probablement openpyxl deprecated warning): {save_error}")
                try:
                    if hasattr(workbook, 'properties'):
                        workbook.properties.modified = None
                    workbook.save(filepath)
                    workbook.close()
                    print(f"Fichier Excel sauvegardé avec succès (tentative alternative): {filepath}")
                except Exception as alt_error:
                    print(f"Erreur lors de la sauvegarde alternative: {alt_error}")
                    try:
                        workbook.close()
                    except Exception:
                        pass
                    raise save_error
            
            print(f"Fichier Excel généré: {filepath}")
            print(f"Template utilisé avec feuilles DQE PGC ({len(df_dqe_pgc)} lignes) et REDEVANCE ({len(df_redevance)} lignes)")
            try:
                os.startfile(filepath)
                print("Fichier Excel ouvert automatiquement")
            except Exception as e:
                print(f"Impossible d'ouvrir automatiquement le fichier: {e}")
            
        except Exception as e:
            print(f"Erreur lors de la génération Excel: {str(e)}")
            import traceback
            print(traceback.format_exc())
    
    def _fill_template_redevance(self, sheet, df_redevance):
        """Remplit la feuille REDEVANCE avec les données"""
        try:
            print(f"Remplissage template REDEVANCE avec {len(df_redevance)} lignes")
            start_row = 2  # Par défaut ligne 2
            for row in range(start_row, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col, value=None)
            for col_idx, column_name in enumerate(df_redevance.columns, 1):
                sheet.cell(row=1, column=col_idx, value=column_name)
            for row_idx, (_, row_data) in enumerate(df_redevance.iterrows(), start_row):
                for col_idx, value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            print(f"Template REDEVANCE rempli: {len(df_redevance)} lignes de données")
            
        except Exception as e:
            print(f"Erreur remplissage template REDEVANCE: {e}")

    def _load_organized_layers(self, results, sro, troncon_safe):
        """Chargement organisé par catégories - VERSION PGC"""
        created_layers = []
        categories = {
            "GC - TDR + RAD (hors fourniture tube PEHD ou PVC et chambres)": [],
            "Chambres": [],
            "Poteaux": [],
            "Fourniture des alvéoles": []
        }
        
        print(f"\n=== DÉBUT TRAITEMENT DQE PGC ORGANISÉ - {len(results)} résultats ===")
        
        for i, result in enumerate(results):
            designation = result.designation if hasattr(result, 'designation') else ""
            
            if not designation:
                print(f"[{i+1}] IGNORÉ - Pas de désignation")
                continue
            
            ids = result.ids if hasattr(result, 'ids') else []
            quantite = result.quantite if hasattr(result, 'quantite') else 0
            
            print(f"\n[{i+1}] Traitement: {designation}")
            print(f"    Quantité: {quantite}")
            print(f"    IDs: {len(ids) if ids else 0} éléments")
            
            if not ids or len(ids) == 0:
                print(f"IGNORÉ - Pas d'IDs (en-tête ou section)")
                continue
            
            try:
                quantite_num = float(quantite) if quantite is not None else 0
                if quantite_num <= 0:
                    print(f"IGNORÉ - Quantité nulle ou négative ({quantite_num})")
                    continue
            except (ValueError, TypeError):
                print(f"IGNORÉ - Quantité invalide: {quantite}")
                continue
            
            main_category = None
            designation_lower = designation.lower()
            if any(x in designation_lower for x in ["chambre", "ch ", "regard"]):
                main_category = "Chambres"
                print(f"Détecté comme Chambre: {designation}")
            elif any(x in designation_lower for x in ["poteau", "pt ", "appui"]):
                main_category = "Poteaux"
                print(f"Détecté comme Poteau: {designation}")
            elif any(x in designation_lower for x in ["alvéole", "alveole", "fourniture alvéole", "fourniture alveole", "tube pehd", "tube pvc", "fourniture tube", "pvc ", "pehd ", " pvc", " pehd"]):
                main_category = "Fourniture des alvéoles"
                print(f"Détecté comme Fourniture alvéole/tube: {designation}")
            elif any(x in designation_lower for x in ["tdr", "rad", "gc ", "génie civil", "tranchée", "forage", "tirage", "pose", "infrastructure", "encorbellement"]):
                main_category = "GC - TDR + RAD (hors fourniture tube PEHD ou PVC et chambres)"
                print(f"Détecté comme GC-TDR+RAD: {designation}")
            
            if not main_category:
                print(f"IGNORÉ - Aucune catégorie déterminée")
                continue
            
            table_name = LayerManager.get_table_from_designation_pgc(designation)
            print(f"Table: {table_name}")
            
            task_data = {
                'designation': designation,
                'ids': ids,
                'table_name': table_name,
                'layer_id': str(uuid.uuid4()),
                'priority': LayerManager.get_custom_layer_order(designation) if hasattr(LayerManager, 'get_custom_layer_order') else 0,
                'sro': sro
            }
            
            categories[main_category].append(task_data)
            print(f"      Assigné à la catégorie: {main_category}")
        for category_name, tasks in categories.items():
            if not tasks:
                continue
            
            print(f"\n=== CHARGEMENT CATÉGORIE: {category_name} ===")
            print(f"Éléments à charger: {len(tasks)}")
            category_group = LayerManager.create_layer_subgroup(self.layer_group, category_name)
            
            for task in tasks:
                if self._is_cancelled:
                    print("Annulation detectee pendant le chargement des couches")
                    return created_layers
                layer_name = f"{task['designation']}"
                if troncon_safe:
                    layer_name += f" - {troncon_safe}"
                
                print(f"Chargement: {layer_name}")
                
                ids_string = ','.join(map(str, task['ids'])) if task['ids'] else ""
                
                layer = LayerManager.load_layer_direct(
                    layer_name,
                    ids_string,
                    task['table_name'],
                    task['sro']
                )
                
                if layer and layer.isValid():
                    QgsProject.instance().addMapLayer(layer, False)
                    category_group.addLayer(layer)
                    created_layers.append(layer)
                    self.layers_loaded.append(layer)
                    
                    print(f"COUCHE CRÉÉE ({layer.featureCount()} entités)")
                else:
                    print(f"ÉCHEC CRÉATION COUCHE: {layer_name}")
        
        print(f"\n=== RÉSUMÉ CHARGEMENT ORGANISÉ ===")
        for category_name, tasks in categories.items():
            if tasks:
                print(f" {category_name}: {len(tasks)} couches")
        
        print(f" Total: {len(created_layers)} couches créées avec organisation")
        return created_layers
