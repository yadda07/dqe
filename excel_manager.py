"""
DQE Excel Manager
======================================================
Gestion des rapports Excel pour le plugin DQE Chargeur
"""

import os
import shutil
import tempfile
import time
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

try:
    from .dqe_utils import _db_manager, _logger
    MODULES_AVAILABLE = True
except ImportError:
    _db_manager = _logger = None
    MODULES_AVAILABLE = False


class ExcelManager:
    @staticmethod
    def get_template_path(operation_type: str) -> str:
        """Retourne le chemin du template selon le type d'opération"""
        plugin_dir = os.path.dirname(__file__)
        template_files = {
            'PRO': 'template_dqe_pro.xlsx',
            'EXE': 'template_dqe_exe.xlsx', 
            'PGC': 'template_dqe_pgc.xlsx'
        }
        
        # Extract base type (EXE_TE -> EXE, PRO -> PRO)
        base_type = operation_type.upper().split('_')[0]
        template_name = template_files.get(base_type, 'template_dqe_pro.xlsx')
        template_path = os.path.join(plugin_dir, 'files', template_name)
        
        print(f"Template recherché pour {operation_type}: {template_path}")
        
        if not os.path.exists(template_path):
            print(f" Template non trouvé: {template_path}")
            fallback_path = os.path.join(plugin_dir, 'files', 'template_dqe_pro.xlsx')
            if os.path.exists(fallback_path):
                print(f"Utilisation du template fallback: {fallback_path}")
                return fallback_path
            else:
                print(" Aucun template trouvé")
                return None
        
        return template_path
    
    @staticmethod
    def create_excel_report(results: List[Dict], sro: str, operation_type: str, troncon: str = None):
        """Génère le rapport Excel en utilisant le bon template"""
        if not results:
            return None
        
        try:
            print(f"\n=== Génération Excel {operation_type.upper()} ===")
            template_path = ExcelManager.get_template_path(operation_type)
            if not template_path:
                print("Attention : Aucun template disponible, génération basique")
                return ExcelManager._create_basic_excel(results, sro, operation_type, troncon)
            redevance_data = None
            if operation_type.upper() == 'PGC' and results:
                first_result = results[0]
                if hasattr(first_result, 'redevance_data') and first_result.redevance_data:
                    redevance_data = first_result.redevance_data
                    print(f"Données REDEVANCE détectées: {len(redevance_data)} lignes")
                else:
                    print("Aucune donnée REDEVANCE trouvée")
            df = pd.DataFrame(results)
            if 'ids' in df.columns:
                print(f"Conservation des IDs des câbles découpés pour {operation_type}")
                df = df.drop(columns=['ids'])  # Supprimer seulement du DataFrame Excel
            if 'redevance_data' in df.columns:
                df = df.drop(columns=['redevance_data'])
            
            if len(df.columns) >= 3:
                df.columns = ["Désignation", "Unité", "Quantité"]
                
                if "Quantité" in df.columns:
                    df["Quantité"] = pd.to_numeric(df["Quantité"], errors='coerce')
                    df["Quantité"] = df["Quantité"].round().fillna(0).astype(int)
            temp_dir = tempfile.gettempdir()
            sro_safe = sro.replace('/', '_')
            
            if troncon:
                excel_path = os.path.join(temp_dir, f"dqe_{operation_type}_{sro_safe}_{troncon}_{int(time.time())}.xlsx")
            else:
                excel_path = os.path.join(temp_dir, f"dqe_{operation_type}_{sro_safe}_{int(time.time())}.xlsx")
            
            print(f"Fichier de sortie: {excel_path}")
            shutil.copy2(template_path, excel_path)
            print(f"Template copié vers: {excel_path}")
            workbook = load_workbook(excel_path)
            sheet_name = f"DQE {operation_type.upper()}"
            target_sheet = None
            for sheet in workbook.sheetnames:
                if operation_type.upper() in sheet.upper():
                    target_sheet = workbook[sheet]
                    print(f"Feuille trouvée: {sheet}")
                    break
            
            if not target_sheet:
                target_sheet = workbook.active
                print(f"Feuille par défaut utilisée: {target_sheet.title}")
            if operation_type.upper() == 'PGC':
                ExcelManager._fill_pgc_template(target_sheet, df, sro, troncon, workbook, redevance_data)
            else:
                ExcelManager._fill_standard_template(target_sheet, df, operation_type)
            try:
                workbook.save(excel_path)
            except Exception as e:
                print(f"Erreur sauvegarde Excel (tentative alternative): {e}")
                try:
                    workbook.properties.creator = None
                    workbook.properties.lastModifiedBy = None
                    workbook.properties.created = None
                    workbook.properties.modified = None
                    workbook.save(excel_path)
                    print("Sauvegarde alternative réussie")
                except Exception as e2:
                    print(f"Échec sauvegarde alternative: {e2}")
                    raise Exception(f"Impossible de sauvegarder le fichier Excel: {e2}")
            
            workbook.close()
            
            print(f" Rapport Excel généré: {excel_path}")
            ExcelManager._open_excel_file(excel_path)
            
            return excel_path
            
        except Exception as e:
            print(f" Erreur génération Excel: {str(e)}")
            import traceback
            print(traceback.format_exc())
            if _logger:
                _logger.error("Erreur génération Excel", exception=e)
            return None
    
    @staticmethod
    def _fill_standard_template(sheet, df, operation_type):
        """Remplit le template standard avec les données"""
        if df.empty:
            return
        
        # Check base type (EXE_TE -> EXE)
        base_type = operation_type.upper().split('_')[0]
        if base_type == 'EXE':
            return ExcelManager._fill_exe_template(sheet, df)
        else:
            updated_rows = []
            
            print(f"\n INDEXATION PRO DIRECTE : {len(df)} données à traiter")
            
            for i, (_, row) in enumerate(df.iterrows()):
                designation = row['Désignation']
                quantite = row['Quantité']
                target_row = i + 2  # +2 car ligne 1 = header
                if target_row <= sheet.max_row:
                    template_cell_value = sheet.cell(row=target_row, column=1).value
                    sheet.cell(row=target_row, column=3, value=quantite)
                    updated_rows.append(f"{designation} -> ligne {target_row}")
                    
                    print(f"  OK data[{i}] -> ligne {target_row}")
                    print(f"      pgAdmin: '{designation}' (Q: {quantite})")
                    print(f"      Template: '{template_cell_value}'")
                else:
                    print(f"  ERREUR data[{i}] -> ligne {target_row} HORS LIMITE (max: {sheet.max_row})")
                    print(f"      pgAdmin: '{designation}' (Q: {quantite})")
            
            print(f"\nTemplate {operation_type.upper()} rempli: {len(updated_rows)} lignes")
            return updated_rows
    
    @staticmethod
    def _open_excel_file(excel_path: str):
        """Ouvre le fichier Excel avec l'application par défaut"""
        try:
            import subprocess
            import platform
            
            system = platform.system()
            if system == 'Windows':
                os.startfile(excel_path)
            elif system == 'Darwin':
                subprocess.call(['open', excel_path])
            else:
                subprocess.call(['xdg-open', excel_path])
                
        except Exception as e:
            print(f"Impossible d'ouvrir le fichier Excel: {e}")
            if _logger:
                _logger.warning("Impossible d'ouvrir automatiquement le fichier Excel", exception=e)
    
    @staticmethod
    def _find_template_row(sheet, designation: str, start_row: int = 1, end_row: int = None) -> Optional[int]:
        """Trouve la ligne correspondante dans le template avec correspondance intelligente pour BPE et câbles"""
        designation_lower = designation.lower().strip()
        
        print(f" Recherche template pour: '{designation}'")
        for row in range(start_row, min(end_row or sheet.max_row + 1, start_row + 200)):
            cell_value = sheet.cell(row=row, column=1).value
            if not cell_value:
                continue
            
            template_text = str(cell_value).lower().strip()
            if template_text == designation_lower:
                print(f"  OK Correspondance exacte trouvée ligne {row}: '{cell_value}'")
                return row
            if "bpe" in designation_lower and "bpe" in template_text:
                import re
                des_fo = re.search(r'(\d+)\s*fo', designation_lower)
                tem_fo = re.search(r'(\d+)\s*fo', template_text)
                
                if des_fo and tem_fo and des_fo.group(1) == tem_fo.group(1):
                    critical_words = ['conduite', 'façade', 'aérien', 'immeuble']
                    matches_critical = True
                    
                    for word in critical_words:
                        des_has_word = word in designation_lower
                        tem_has_word = word in template_text
                        if des_has_word != tem_has_word:
                            matches_critical = False
                            break
                    
                    if matches_critical:
                        print(f"  OK Correspondance BPE {des_fo.group(1)} FO trouvée ligne {row}: '{cell_value}'")
                        return row
                    else:
                        print(f"  Attention : BPE {des_fo.group(1)} FO trouvée ligne {row} mais mots-clés différents: '{cell_value}'")
            elif "câble" in designation_lower and "câble" in template_text:
                import re
                des_fo = re.search(r'(\d+)\s*fo', designation_lower)
                tem_fo = re.search(r'(\d+)\s*fo', template_text)
                
                if des_fo and tem_fo and des_fo.group(1) == tem_fo.group(1):
                    critical_words = ['conduite', 'façade', 'aérien', 'immeuble']
                    matches_critical = True
                    
                    for word in critical_words:
                        des_has_word = word in designation_lower
                        tem_has_word = word in template_text
                        if des_has_word != tem_has_word:
                            matches_critical = False
                            break
                    
                    if matches_critical:
                        print(f"  OK Correspondance câble {des_fo.group(1)} FO trouvée ligne {row}: '{cell_value}'")
                        return row
                    else:
                        print(f"  Attention : Câble {des_fo.group(1)} FO trouvée ligne {row} mais mots-clés différents: '{cell_value}'")
            elif len(designation_lower) > 5 and len(template_text) > 5:
                from difflib import SequenceMatcher
                similarity = SequenceMatcher(None, designation_lower, template_text).ratio()
                if similarity >= 0.85:  # 85% de similarité pour les autres éléments
                    print(f"  OK Correspondance par similarité ({similarity:.2f}) trouvée ligne {row}: '{cell_value}'")
                    return row
        
        print(f" Aucune correspondance trouvée pour '{designation}'")
        return None

    @staticmethod
    def _fill_exe_template(sheet, df):
        """Remplit le template EXE (indexation directe + alvéoles dynamiques)"""
        if df.empty:
            return []
            
        updated_rows = []
        alveoles_data = []
        non_alveole_data = []
        
        print(f"\n INDEXATION EXE DIRECTE : {len(df)} données à traiter")
        for i, (_, row) in enumerate(df.iterrows()):
            designation = row['Désignation']
            quantite = row['Quantité']
            unite = row.get('Unité', 'ml')
            designation_lower = designation.lower()
            is_alveole = False
            if any(keyword in designation_lower for keyword in ['pvc ', 'pehd ']):
                import re
                if re.search(r'pvc\s+\d+|pehd\s+\d+|\(.*fois.*\)', designation_lower):
                    if not any(exclude in designation_lower for exclude in ['gc -', 'tdr', 'rad', 'génie civil']):
                        is_alveole = True
            elif any(keyword in designation_lower for keyword in ['alvéole', 'alveole']):
                if not any(exclude in designation_lower for exclude in ['gc -', 'tdr', 'rad', 'génie civil', 'fourniture des']):
                    is_alveole = True
            
            if is_alveole:
                alveoles_data.append({
                    'designation': designation,
                    'unite': unite,
                    'quantite': quantite
                })
                print(f"  → ALVÉOLE détectée: {designation}")
            elif 'fourniture des alvéoles' in designation_lower:
                print(f"  → TITRE SECTION ignoré: {designation}")
            else:
                non_alveole_data.append({
                    'designation': designation,
                    'unite': unite,
                    'quantite': quantite,
                    'original_index': i
                })
        print(f"  → Traitement de {len(non_alveole_data)} éléments non-alvéoles par indexation directe")
        
        for i, data in enumerate(non_alveole_data):
            designation = data['designation']
            quantite = data['quantite']
            target_row = i + 2  # +2 car ligne 1 = header
            if target_row <= sheet.max_row:
                template_cell_value = sheet.cell(row=target_row, column=1).value
                sheet.cell(row=target_row, column=3, value=quantite)
                updated_rows.append(f"{designation} -> ligne {target_row}")
                
                print(f"  OK data[{i}] -> ligne {target_row}")
                print(f"      pgAdmin: '{designation}' (Q: {quantite})")
                print(f"      Template: '{template_cell_value}'")
            else:
                print(f"  ERREUR data[{i}] -> ligne {target_row} HORS LIMITE (max: {sheet.max_row})")
                print(f"      pgAdmin: '{designation}' (Q: {quantite})")
        if alveoles_data:
            print(f"  → Traitement de {len(alveoles_data)} alvéoles dynamiques...")
            alveoles_header_row = None
            for row in range(130, min(sheet.max_row + 10, 150)):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and 'fourniture des alvéoles' in str(cell_value).lower():
                    alveoles_header_row = row
                    print(f"  → Section alvéoles trouvée ligne {row}")
                    break
            
            if alveoles_header_row:
                next_row = alveoles_header_row + 1
                for alv_data in alveoles_data:
                    while next_row <= sheet.max_row and sheet.cell(row=next_row, column=1).value:
                        next_row += 1
                    
                    sheet.cell(row=next_row, column=1, value=alv_data['designation'])
                    sheet.cell(row=next_row, column=2, value=alv_data['unite'])
                    sheet.cell(row=next_row, column=3, value=alv_data['quantite'])
                    updated_rows.append(f"ALVÉOLE {alv_data['designation']} -> ligne {next_row}")
                    print(f"  OK ALVÉOLE: {alv_data['designation']} -> ligne {next_row}")
                    next_row += 1
            else:
                print(f"  Attention : Section alvéoles non trouvée, alvéoles ignorées")
        
        print(f"\nTemplate EXE rempli: {len(updated_rows)} lignes")
        return updated_rows

    @staticmethod
    def _fill_pgc_template(sheet, df: pd.DataFrame, sro: str, troncon: str, workbook=None, redevance_data: List[Dict] = None):
        """Remplit spécifiquement le template PGC avec correspondance exacte par désignation"""
        print(f"\n=== REMPLISSAGE TEMPLATE PGC (CORRESPONDANCE EXACTE) ===")
        print(f"Données: {len(df)} lignes")
        print(f"SRO: {sro}, Tronçon: {troncon}")
        for row in range(1, min(20, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and "Nom GC :" in str(cell_value):
                new_gc_text = f"Nom GC : {troncon}"
                sheet.cell(row=row, column=1, value=new_gc_text)
                print(f"Ligne GC mise a jour: '{new_gc_text}'")
                break
        matched_count = 0
        alveoles_added = []
        
        for _, result_row in df.iterrows():
            sql_designation = str(result_row['Désignation']).strip()
            sql_unite = str(result_row.get('Unité', '')).strip() 
            sql_quantite = result_row['Quantité']
            if not sql_designation or sql_designation in ['', 'Désignation', 'Aucune donnée']:
                continue
            if any(x in sql_designation.lower() for x in ["pvc ", "pehd"]):
                alveoles_added.append({
                    'designation': sql_designation,
                    'unite': sql_unite, 
                    'quantite': sql_quantite
                })
                continue
            template_row = None
            for row in range(1, sheet.max_row + 1):
                template_cell = sheet.cell(row=row, column=1).value
                if template_cell and str(template_cell).strip().lower() == sql_designation.lower():
                    template_row = row
                    break
            if template_row:
                try:
                    quantite_num = float(sql_quantite) if sql_quantite is not None else 0
                    sheet.cell(row=template_row, column=3, value=quantite_num)
                    print(f"EXACT: {sql_designation} -> ligne {template_row} = {quantite_num}")
                    matched_count += 1
                except (ValueError, TypeError):
                    print(f"WARN: Quantite invalide pour {sql_designation}: {sql_quantite}")
            else:
                print(f"WARN: Aucune correspondance pour {sql_designation}")
        if alveoles_added:
            alveoles_section_row = None
            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and "Fourniture des Alvéoles" in str(cell_value):
                    alveoles_section_row = row
                    break
            
            if alveoles_section_row:
                next_row = alveoles_section_row + 1
                for alv_data in alveoles_added:
                    while next_row <= sheet.max_row and sheet.cell(row=next_row, column=1).value:
                        next_row += 1
                    
                    sheet.cell(row=next_row, column=1, value=alv_data['designation'])
                    sheet.cell(row=next_row, column=2, value=alv_data['unite'])
                    try:
                        quantite_num = float(alv_data['quantite']) if alv_data['quantite'] is not None else 0
                        sheet.cell(row=next_row, column=3, value=quantite_num)
                        print(f"ALVEOLE: {alv_data['designation']} -> ligne {next_row} = {quantite_num}")
                        matched_count += 1
                    except (ValueError, TypeError):
                        pass
                    next_row += 1
        if redevance_data and workbook:
            print(f"\n=== REMPLISSAGE FEUILLE REDEVANCE ===")
            ExcelManager._fill_redevance_sheet(workbook, redevance_data, sro, troncon)
        
        print(f"Template DQE PGC: {matched_count} éléments remplis, {len(alveoles_added)} alvéoles ajoutées")
    
    @staticmethod
    def _find_next_section(sheet, current_section_row: int) -> int:
        """Trouve la ligne de fin de la section courante (début de la section suivante)"""
        for row in range(current_section_row + 1, min(current_section_row + 50, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str):
                cell = sheet.cell(row=row, column=1)
                if (cell.font and cell.font.bold) or any(keyword in cell_value.lower() for keyword in [
                    "fourniture", "pose de", "armoire", "gc -", "transport", "distribution"
                ]):
                    return row - 1
        return current_section_row + 20
    
    @staticmethod
    def _find_template_row(sheet, designation: str, start_row: int = 1, end_row: int = None) -> Optional[int]:
        """Trouve la ligne correspondante dans le template avec correspondance intelligente pour BPE et câbles"""
        designation_lower = designation.lower().strip()
        
        print(f" Recherche template pour: '{designation}'")
        for row in range(start_row, min(end_row or sheet.max_row + 1, start_row + 200)):
            cell_value = sheet.cell(row=row, column=1).value
            if not cell_value:
                continue
            
            template_text = str(cell_value).lower().strip()
            if template_text == designation_lower:
                print(f"  OK Correspondance exacte trouvée ligne {row}: '{cell_value}'")
                return row
            if "bpe" in designation_lower and "bpe" in template_text:
                import re
                des_fo = re.search(r'(\d+)\s*fo', designation_lower)
                tem_fo = re.search(r'(\d+)\s*fo', template_text)
                
                if des_fo and tem_fo and des_fo.group(1) == tem_fo.group(1):
                    critical_words = ['conduite', 'façade', 'aérien', 'immeuble']
                    matches_critical = True
                    
                    for word in critical_words:
                        des_has_word = word in designation_lower
                        tem_has_word = word in template_text
                        if des_has_word != tem_has_word:
                            matches_critical = False
                            break
                    
                    if matches_critical:
                        print(f"  OK Correspondance BPE {des_fo.group(1)} FO trouvée ligne {row}: '{cell_value}'")
                        return row
                    else:
                        print(f"  Attention : BPE {des_fo.group(1)} FO trouvée ligne {row} mais mots-clés différents: '{cell_value}'")
            elif "câble" in designation_lower and "câble" in template_text:
                import re
                des_fo = re.search(r'(\d+)\s*fo', designation_lower)
                tem_fo = re.search(r'(\d+)\s*fo', template_text)
                
                if des_fo and tem_fo and des_fo.group(1) == tem_fo.group(1):
                    critical_words = ['conduite', 'façade', 'aérien', 'immeuble']
                    matches_critical = True
                    
                    for word in critical_words:
                        des_has_word = word in designation_lower
                        tem_has_word = word in template_text
                        if des_has_word != tem_has_word:
                            matches_critical = False
                            break
                    
                    if matches_critical:
                        print(f"  OK Correspondance câble {des_fo.group(1)} FO trouvée ligne {row}: '{cell_value}'")
                        return row
                    else:
                        print(f"  Attention : Câble {des_fo.group(1)} FO trouvée ligne {row} mais mots-clés différents: '{cell_value}'")
            elif len(designation_lower) > 5 and len(template_text) > 5:
                designation_clean = ExcelManager._smart_match(designation)
                template_clean = ExcelManager._smart_match(str(cell_value))
                
                from difflib import SequenceMatcher
                similarity = SequenceMatcher(None, designation_clean, template_clean).ratio()
                if similarity >= 0.85:  # 85% de similarité pour les autres éléments
                    print(f"  OK Correspondance par similarité ({similarity:.2f}) trouvée ligne {row}: '{cell_value}'")
                    return row
        
        print(f" Aucune correspondance trouvée pour '{designation}'")
        return None
    
    @staticmethod
    def _smart_match(text: str) -> str:
        """Nettoie et normalise le texte pour la correspondance intelligente"""
        import re
        
        if not text:
            return ""
        text = text.lower().strip()
        text = re.sub(r'(\d+)\s*fo\b', r'\1 fo', text)
        if any(keyword in text for keyword in ['pvc', 'pehd', 'alvéole', 'alveole']):
            pass
        else:
            text = re.sub(r'\s*[\(\[\{].*$', '', text)
        text = re.sub(r'câble\s+optique', 'câble', text)
        text = re.sub(r'[^\w\s\(\)\/\-\.]', '', text)
        text = re.sub(r'\s+', ' ', text).strip()
        
        return text
    
    @staticmethod
    def _create_basic_excel(results: List[Dict], sro: str, operation_type: str, troncon: str = None):
        """Génération Excel basique si pas de template"""
        try:
            df = pd.DataFrame(results)
            
            if 'ids' in df.columns:
                df = df.drop(columns=['ids'])
            
            if len(df.columns) >= 3:
                df.columns = ["Désignation", "Unité", "Quantité"]
                
                if "Quantité" in df.columns:
                    df["Quantité"] = pd.to_numeric(df["Quantité"], errors='coerce')
                    df["Quantité"] = df["Quantité"].round().fillna(0).astype(int)
            
            temp_dir = tempfile.gettempdir()
            sro_safe = sro.replace('/', '_')
            
            if troncon:
                excel_path = os.path.join(temp_dir, f"dqe_{operation_type}_basic_{sro_safe}_{troncon}_{int(time.time())}.xlsx")
            else:
                excel_path = os.path.join(temp_dir, f"dqe_{operation_type}_basic_{sro_safe}_{int(time.time())}.xlsx")
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=f'DQE {operation_type.upper()}', index=False)
            
            return excel_path
            
        except Exception as e:
            print(f"Erreur génération Excel basique: {e}")
            return None
    
    @staticmethod
    def _open_excel_file(excel_path: str):
        """Ouvre le fichier Excel avec l'application par défaut"""
        try:
            import subprocess
            import platform
            
            system = platform.system()
            if system == 'Windows':
                os.startfile(excel_path)
            elif system == 'Darwin':
                subprocess.call(['open', excel_path])
            else:
                subprocess.call(['xdg-open', excel_path])
                
        except Exception as e:
            print(f"Impossible d'ouvrir le fichier Excel: {e}")
            if _logger:
                _logger.warning("Impossible d'ouvrir automatiquement le fichier Excel", exception=e)

    @staticmethod
    def _fill_redevance_sheet(workbook, redevance_data: List[Dict], sro: str, troncon: str):
        """Remplit la feuille REDEVANCE avec les données de gc_exe.redevance_table"""
        print(f"Remplissage feuille REDEVANCE avec {len(redevance_data)} lignes")
        
        try:
            if 'REDEVANCE' in workbook.sheetnames:
                redevance_sheet = workbook['REDEVANCE']
                print("Feuille REDEVANCE existante trouvee - preservation des en-tetes")
                start_row = 2  # Par défaut ligne 2 (ligne 1 = en-têtes)
                existing_headers = []
                for col in range(1, redevance_sheet.max_column + 1):
                    cell_value = redevance_sheet.cell(row=1, column=col).value
                    if cell_value:
                        existing_headers.append(str(cell_value))
                
                if existing_headers:
                    print(f"En-tetes existants trouves: {existing_headers}")
                    print("Les en-tetes seront preserves, remplissage des donnees seulement")
                else:
                    print("Attention : Aucun en-tête existant trouvé")
            
            else:
                print("Attention : Feuille REDEVANCE non trouvée dans le template, création d'une nouvelle feuille")
                redevance_sheet = workbook.create_sheet(title='REDEVANCE')
                start_row = 2  # Ligne 2 pour les données (ligne 1 pour en-têtes)
                existing_headers = []
            max_row = redevance_sheet.max_row
            if max_row > 1:
                for row in redevance_sheet.iter_rows(min_row=2, max_row=max_row):
                    for cell in row:
                        cell.value = None
                print(f"Donnees precedentes effacees (lignes 2-{max_row})")
            if not redevance_data:
                print("Attention : Aucune donnée redevance à remplir")
                return
            df_redevance = pd.DataFrame(redevance_data)
            print(f"Colonnes des données: {list(df_redevance.columns)}")
            if not existing_headers:
                print("Ecriture des en-tetes (feuille nouvelle ou vide)")
                for col_idx, column_name in enumerate(df_redevance.columns, start=1):
                    redevance_sheet.cell(row=1, column=col_idx, value=column_name)
            else:
                print("En-tetes preserves (feuille existante)")
            for row_idx, (_, row_data) in enumerate(df_redevance.iterrows(), start=start_row):
                for col_idx, value in enumerate(row_data, start=1):
                    if pd.isna(value):
                        excel_value = None
                    elif isinstance(value, (int, float)):
                        excel_value = float(value) if value != int(value) else int(value)
                    else:
                        excel_value = str(value)
                    
                    redevance_sheet.cell(row=row_idx, column=col_idx, value=excel_value)
            for column in redevance_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max 50 caractères
                redevance_sheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Feuille REDEVANCE remplie avec {len(df_redevance)} lignes de donnees")
            print(f"Colonnes: {', '.join(df_redevance.columns)}")
            
        except Exception as e:
            print(f"CRITICAL Erreur lors du remplissage de la feuille REDEVANCE: {e}")
            if _logger:
                _logger.error("Erreur remplissage feuille REDEVANCE", exception=e)
