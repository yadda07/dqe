"""
DQE PRO Tab Module

Handles the DQE PRO tab interface and functionality for the QGIS plugin.
Extracted from the main dialog file for better modularity.
"""

import os
import json
import time
import uuid

from .compat import (
    QTimer, QVBoxLayout, QHBoxLayout, QFormLayout, QGroupBox,
    QPushButton, QComboBox, QMessageBox, QApplication, QGIS_SUCCESS
)
from .base_tab import BaseDQETab
from qgis.utils import iface
from .ui_components import SROComboBox, ProgressWidget
from .layer_manager import LayerManager
from .database_operations import DatabaseOperations
from .excel_manager import ExcelManager
from .dqe_utils import _db_manager, _logger, _validator
from .designation_classifier import DesignationClassifier
from .workers import DQEWorker, DistCablesWorker
from .compat import QThread
from .dqe_utils import _crash_log
from .telemetry import send_telemetry


class DQEProTab(BaseDQETab):
    TAB_LABEL = "PRO"
    """Interface et logique pour l'onglet DQE PRO"""
    
    @staticmethod
    def filter_results_by_template(results):
        """Filtre les résultats DQE Pro pour ne garder que les lignes présentes dans le template Excel"""
        try:
            plugin_dir = os.path.dirname(__file__)
            template_path = os.path.join(plugin_dir, 'files', 'template_dqe_pro.xlsx')
            
            if not os.path.exists(template_path):
                if _logger:
                    _logger.warning(f"Template Excel non trouve: {template_path}")
                return results  # Retourner tous les résultats si pas de template
            
            # Lecture du template Excel
            try:
                import openpyxl
                wb = openpyxl.load_workbook(template_path, read_only=True)
                sheet = wb.active
                
                template_designations = set()
                for row in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=1).value
                    if cell_value and isinstance(cell_value, str):
                        designation = cell_value.strip()
                        if designation and designation != "Désignation":
                            template_designations.add(designation)
                
                wb.close()
                if _logger:
                    _logger.debug(f"Template Excel charge: {len(template_designations)} designations")
                
            except ImportError:
                if _logger:
                    _logger.warning("openpyxl non disponible, pas de filtrage")
                return results
            
            filtered_results = []
            excluded_count = 0
            
            for result in results:
                designation = result.get("designation") or result.get("Désignation") or result.get("désignation") or ""
                
                if designation in template_designations:
                    filtered_results.append(result)
                else:
                    excluded_count += 1
                    pass
            
            if _logger:
                _logger.info(f"Filtrage termine: {len(filtered_results)} conservees, {excluded_count} exclues")
            return filtered_results
            
        except Exception as e:
            if _logger:
                _logger.error(f"Erreur lors du filtrage: {str(e)}")
            return results  # En cas d'erreur, retourner tous les résultats

    def __init__(self, parent=None):
        super().__init__(parent)
        self.dqe_results = []
        self.current_sro = None
        self.current_type = None
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(4)
        
        config_group = QGroupBox("Configuration DQE PRO")
        config_layout = QFormLayout(config_group)
        
        self.sro_input = SROComboBox()
        config_layout.addRow("SRO:", self.sro_input)
        
        self.type_combo = QComboBox()
        self.type_combo.addItem("Transport", "T")
        self.type_combo.addItem("Distribution", "D")
        config_layout.addRow("Type:", self.type_combo)
        
        layout.addWidget(config_group)
        
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(8)  # Espacement réduit entre boutons
        
        self.execute_button = QPushButton("Exécuter DQE PRO")
        self.execute_button.setToolTip("Génère le DQE et charge les couches QGIS")
        self.execute_button.clicked.connect(self.execute_dqe_pro)
        buttons_layout.addWidget(self.execute_button)
        
        self.validate_button = QPushButton("Valider DQE")
        self.validate_button.setToolTip("Enregistre le DQE dans la base de données")
        self.validate_button.clicked.connect(self.validate_dqe_pro)
        buttons_layout.addWidget(self.validate_button)
        
        layout.addLayout(buttons_layout)
        
        self.progress_widget = ProgressWidget()
        layout.addWidget(self.progress_widget)
        
        self.execute_button.setEnabled(False)
        self.sro_input.lineEdit().textChanged.connect(self.on_sro_changed)
    
    def on_sro_changed(self):
        """Valide le SRO de maniere asynchrone apres un delai de saisie"""
        sro = self.sro_input.lineEdit().text().strip()
        self.execute_button.setEnabled(False)
        if sro and len(sro) >= 3:
            QTimer.singleShot(1000, lambda: self.validate_sro_async(sro))
    
    def validate_sro_async(self, sro: str):
        """Validation format uniquement sur frappe (pas de DB).

        La validation d'existence en base est différée au clic sur
        execute_dqe_pro pour éviter tout appel DB sur le main thread
        pendant la saisie.
        """
        if sro != self.sro_input.lineEdit().text().strip():
            return
        if _validator:
            if _validator._validate_sro_format(sro):
                self.execute_button.setEnabled(True)
    
    def execute_dqe_pro(self):
        sro = self.sro_input.lineEdit().text().strip()
        p_type = self.type_combo.currentData()
        _crash_log.step("execute_dqe_pro START", f"sro={sro} type={p_type}")
        
        if not sro:
            QMessageBox.warning(self, "Erreur", "Veuillez saisir un SRO")
            return
        
        if _validator:
            is_valid, message = _validator.validate_sro_exists(sro)
            if not is_valid:
                QMessageBox.warning(self, "SRO invalide", f"Le SRO saisi n'est pas valide:\n{message}")
                return
        
        self.execute_button.setEnabled(False)
        self._exec_start = time.monotonic()
        
        try:
            worker = DQEWorker("PRO", sro, p_type)
            _crash_log.step("execute_dqe_pro", "calling _setup_worker_thread")
            self._setup_worker_thread(worker, self.on_dqe_pro_finished)
            
        except Exception as e:
            error_msg = f"Erreur initialisation DQE PRO: {str(e)}"
            if _logger:
                _logger.error(error_msg)
            self.progress_widget.complete_operation(False, error_msg)
            self.execute_button.setEnabled(True)
            QMessageBox.critical(self, "Erreur", error_msg)
    
    def on_dqe_pro_finished(self, success: bool, results, message: str):
        """Callback appelé quand le traitement DQE PRO est terminé"""
        _crash_log.step("on_dqe_pro_finished START", f"success={success} results={len(results) if results else 0}")
        try:
            if hasattr(self, 'progress_timer'):
                self.post_processing = True
                
            if success and results:
                self.smooth_progress_to(92, "Création des couches...")
                QApplication.processEvents()
                current_date = time.strftime("%Y-%m-%d_%H%M%S")
                sro = self.sro_input.lineEdit().text().strip()
                sro_safe = sro.replace('/', '_')
                group_name = f"DQE_PRO_{sro_safe}_{current_date}"
                self.layer_group = LayerManager.create_layer_group(group_name)
                
                # Stockage des résultats SQL pour validation ultérieure
                self.dqe_results = results
                self.current_sro = sro
                self.current_type = self.type_combo.currentData()
                if _logger:
                    _logger.debug(f"{len(results)} resultats SQL stockes pour validation")
                
                _crash_log.step("on_dqe_pro_finished", "calling _load_organized_layers")
                created_layers = self._load_organized_layers(results, sro, self.type_combo.currentData())
                _crash_log.step("on_dqe_pro_finished", f"layers created={len(created_layers)}")
                
                if self._is_cancelled:
                    self._cleanup_orphan_layers()
                    self.progress_widget.complete_operation(False, "Operation annulee")
                    send_telemetry(
                        action="execution", mode="PRO",
                        sro=sro, type=self.type_combo.currentData(),
                        projet_code="TP" if self.type_combo.currentData() == "T" else "DP",
                        verdict="cancelled", cancelled=True,
                        elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
                    )
                    return
                
                # Stocker pour la phase finale (Excel, collapse)
                self._pending_created_layers = created_layers
                self._pending_results = results
                self._pending_sro = sro
                
                QApplication.processEvents()
                if self.type_combo.currentData() == 'D' and not self._is_cancelled:
                    _crash_log.step("on_dqe_pro_finished", "Distribution: launching dist_cables_worker")
                    if _logger:
                        _logger.info("Lancement worker async cables decoupes (Distribution)")
                    self._start_dist_cables_worker(sro)
                    return  # La suite se fait dans on_dist_cables_finished
                
                self._finalize_dqe_pro()
            else:
                self.progress_widget.complete_operation(False, message)
                QMessageBox.critical(self, "Erreur", message)
                send_telemetry(
                    action="execution", mode="PRO",
                    sro=self.sro_input.lineEdit().text().strip(),
                    type=self.type_combo.currentData(),
                    projet_code="TP" if self.type_combo.currentData() == "T" else "DP",
                    verdict="failure", error_msg=message,
                    elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
                )
                
        except Exception as e:
            error_msg = f"Erreur post-traitement DQE PRO: {str(e)}"
            import traceback
            if _logger:
                _logger.error(error_msg)
                _logger.error(traceback.format_exc())
            self.progress_widget.complete_operation(False, error_msg)
            QMessageBox.critical(self, "Erreur", error_msg)
            send_telemetry(
                action="execution", mode="PRO",
                sro=self.sro_input.lineEdit().text().strip(),
                type=self.type_combo.currentData(),
                projet_code="TP" if self.type_combo.currentData() == "T" else "DP",
                verdict="failure", error_msg=error_msg,
                elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
            )
        finally:
            self._cleanup_thread()
            self.execute_button.setEnabled(True)
    
    def _start_dist_cables_worker(self, sro):
        """Lance le worker async pour la phase DB des cables decoupes."""
        _crash_log.step("_start_dist_cables_worker START", f"sro={sro}")
        self._dist_worker = DistCablesWorker(sro)
        self._dist_thread = QThread()
        _crash_log.step("_start_dist_cables_worker", "moveToThread")
        self._dist_worker.moveToThread(self._dist_thread)
        self._dist_worker.finished.connect(self._on_dist_cables_finished)
        self.progress_widget.progress_cancelled.connect(self._dist_worker.cancel)
        self._dist_thread.started.connect(self._dist_worker.run)
        self._dist_thread.finished.connect(self._dist_thread.deleteLater)
        _crash_log.step("_start_dist_cables_worker", "thread.start()")
        self._dist_thread.start()
        _crash_log.step("_start_dist_cables_worker END")
    
    def _on_dist_cables_finished(self, success, db_result, message):
        """Callback apres phase DB async des cables decoupes."""
        _crash_log.step("_on_dist_cables_finished START", f"success={success}")
        try:
            if self._dist_worker:
                try:
                    self._dist_worker.deleteLater()
                except RuntimeError:
                    pass
                self._dist_worker = None
            if self._dist_thread:
                try:
                    self._dist_thread.quit()
                    self._dist_thread.wait(5000)
                except RuntimeError:
                    pass
                self._dist_thread = None
            
            if self._is_cancelled:
                self._cleanup_orphan_layers()
                self.progress_widget.complete_operation(False, "Operation annulee")
                send_telemetry(
                    action="execution", mode="PRO",
                    sro=self.sro_input.lineEdit().text().strip(),
                    type=self.type_combo.currentData(),
                    projet_code="TP" if self.type_combo.currentData() == "T" else "DP",
                    verdict="cancelled", cancelled=True,
                    elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
                )
                self._cleanup_thread()
                self.execute_button.setEnabled(True)
                return
            
            if success and db_result:
                _crash_log.step("_on_dist_cables_finished", "calling create_distribution_layers")
                self.smooth_progress_to(95, "Creation couches cables decoupes...")
                QApplication.processEvents()
                dist_layers = LayerManager.create_distribution_layers(
                    db_result, self.layer_group, self.layers_loaded,
                    cancel_check=lambda: self._is_cancelled
                )
                _crash_log.step("_on_dist_cables_finished", f"dist_layers={len(dist_layers)}")
                if self._is_cancelled:
                    self._cleanup_orphan_layers()
                    self.progress_widget.complete_operation(False, "Operation annulee")
                    self._cleanup_thread()
                    self.execute_button.setEnabled(True)
                    return
                self._pending_created_layers.extend(dist_layers)
                if _logger:
                    _logger.info(f"Cables decoupes: {len(dist_layers)} couches creees")
            elif not success and message:
                if _logger:
                    _logger.warning(f"Cables decoupes: {message}")
            
            self._finalize_dqe_pro()
            
        except Exception as e:
            error_msg = f"Erreur cables decoupes: {e}"
            if _logger:
                import traceback
                _logger.error(error_msg)
                _logger.error(traceback.format_exc())
            self.progress_widget.complete_operation(False, error_msg)
            self._cleanup_thread()
            self.execute_button.setEnabled(True)
    
    def _finalize_dqe_pro(self):
        """Phase finale commune : Excel, collapse, message."""
        _crash_log.step("_finalize_dqe_pro START")
        try:
            if self._is_cancelled:
                self._cleanup_orphan_layers()
                self.progress_widget.complete_operation(False, "Operation annulee")
                send_telemetry(
                    action="execution", mode="PRO",
                    sro=self._pending_sro if hasattr(self, '_pending_sro') else "",
                    type=self.type_combo.currentData() if hasattr(self, 'type_combo') else "",
                    projet_code="TP" if self.type_combo.currentData() == "T" else "DP",
                    verdict="cancelled", cancelled=True,
                    elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
                )
                return
            
            sro = self._pending_sro
            results = self._pending_results
            created_layers = self._pending_created_layers
            
            self.smooth_progress_to(98, "Generation du rapport Excel...")
            QApplication.processEvents()
            ExcelManager.create_excel_report(results, sro, "PRO")
            self.smooth_progress_to(100, "Finalisation...")
            QApplication.processEvents()
            
            if self.layer_group:
                LayerManager.collapse_group_recursive(self.layer_group)
            
            final_message = f"DQE PRO termine: {len(created_layers)} couches creees"
            self.progress_widget.complete_operation(True, final_message)
            send_telemetry(
                action="execution", mode="PRO",
                sro=sro, type=self.type_combo.currentData(),
                projet_code="TP" if self.type_combo.currentData() == "T" else "DP",
                verdict="success",
                n_results=len(results) if results else 0,
                n_layers=len(created_layers) if created_layers else 0,
                excel_generated=True,
                elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
            )
            
            if iface:
                iface.messageBar().pushMessage(
                    "DQE PRO", final_message,
                    level=QGIS_SUCCESS, duration=5
                )
        except Exception as e:
            error_msg = f"Erreur finalisation DQE PRO: {e}"
            if _logger:
                _logger.error(error_msg)
            self.progress_widget.complete_operation(False, error_msg)
            send_telemetry(
                action="execution", mode="PRO",
                sro=self._pending_sro if hasattr(self, '_pending_sro') else "",
                type=self.type_combo.currentData() if hasattr(self, 'type_combo') else "",
                projet_code="TP" if self.type_combo.currentData() == "T" else "DP",
                verdict="failure", error_msg=error_msg,
                elapsed_ms=int((time.monotonic() - getattr(self, '_exec_start', time.monotonic())) * 1000),
            )
        finally:
            self.execute_button.setEnabled(True)
    
    def _load_organized_layers(self, results, sro, p_type):
        """Chargement organisé par catégories"""
        created_layers = []
        categories = DesignationClassifier.get_categories("PRO")
        
        if _logger:
            _logger.info(f"Debut traitement DQE PRO - {len(results)} resultats")
        
        for i, result in enumerate(results):
            designation = result.get("designation") or result.get("Désignation") or result.get("désignation") or ""
            if not designation:
                continue
            
            ids = result.get("ids") or result.get("Ids") or ""
            quantite = result.get("quantite") or result.get("Quantité") or result.get("quantité") or 0
            
            if not ids or len(str(ids).strip()) == 0:
                continue
            
            try:
                quantite_num = float(quantite) if quantite is not None else 0
                if quantite_num <= 0:
                    continue
            except (ValueError, TypeError):
                continue
            
            main_category = DesignationClassifier.classify(designation, "PRO", p_type)
            if main_category is None:
                if _logger:
                    _logger.debug(f"Ignore - cable remplace par cables decoupes: {designation}")
                continue
            
            table_name = LayerManager.get_table_from_designation(designation)
            task_data = {
                'layer_name': designation,
                'ids_str': str(ids),
                'table_name': table_name,
                'layer_id': str(uuid.uuid4()),
                'priority': LayerManager.get_custom_layer_order(designation),
                'sro': sro
            }
            categories[main_category].append(task_data)
            if _logger:
                _logger.debug(f"[{i+1}] {designation} -> {main_category}")

        created_layers = self._load_categories_to_layers(categories, sro)
        return created_layers
    
    def validate_dqe_pro(self):
        """Valide et sauvegarde le DQE PRO dans la base de données
        Structure simplifiée: 1 ligne avec categorie='dqe_result' et tout le DQE en JSON
        """
        validate_start = time.monotonic()
        try:
            sro = self.sro_input.currentText().strip()
            if not sro:
                QMessageBox.warning(self, "Validation DQE", "Veuillez sélectionner un SRO")
                return
            
            if not self.dqe_results:
                QMessageBox.warning(self, "Validation DQE", 
                    "Aucun résultat DQE trouvé.\nVeuillez d'abord exécuter le DQE PRO.")
                return
            
            type_data = self.type_combo.currentData()
            projet_code = "TP" if type_data == "T" else "DP"
            user_name = _db_manager.config.user if _db_manager and _db_manager.config else "unknown"
            
            dqe_data = self._build_dqe_data(self.dqe_results)
            
            if not dqe_data:
                QMessageBox.warning(self, "Validation DQE", "Aucune donnée DQE")
                return
            
            if _logger:
                _logger.info(f"Validation DQE PRO: {len(dqe_data)} lignes a enregistrer")
            
            # Une seule insertion avec tout le DQE
            with _db_manager.get_cursor() as cursor:
                query = """
                    INSERT INTO dqe.dqejson 
                    (sro, nom_dqe, projet, categorie, champs, user_name, version_projet) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    sro,
                    f"DQE_PRO_{sro}",
                    projet_code,
                    'dqe_result',  # Catégorie unique
                    json.dumps(dqe_data),  # Tableau JSON complet
                    user_name,
                    None  # Version auto-assignée par trigger
                ))
            
            if _logger:
                _logger.info(f"Validation: 1 ligne dqe_result avec {len(dqe_data)} elements")
            
            # Sauvegarde de TOUTES les couches avec géométries
            layers_count = 0
            if self.layer_group:
                layers_count = self._save_all_layers(sro, projet_code, user_name)
            
            total_saved = len(dqe_data) + layers_count
            if _logger:
                _logger.info(f"Validation terminee: {len(dqe_data)} resultats SQL + {layers_count} couches")
            
            QMessageBox.information(
                self, 
                "Validation DQE PRO", 
                f"Validation terminee!\n\n"
                f"- SRO: {sro}\n"
                f"- Type: {projet_code}\n"
                f"- Resultats DQE: {len(dqe_data)}\n"
                f"- Couches archivees: {layers_count}\n"
                f"- Total: {total_saved} elements"
            )
            send_telemetry(
                action="validation", mode="PRO",
                sro=sro, type=type_data, projet_code=projet_code,
                verdict="success",
                n_dqe_data=len(dqe_data), n_layers_saved=layers_count,
                elapsed_ms=int((time.monotonic() - validate_start) * 1000),
            )
            
        except Exception as e:
            import traceback
            if _logger:
                _logger.error(f"Erreur validation: {str(e)}")
                _logger.error(traceback.format_exc())
            if _logger:
                _logger.error("Erreur validation DQE PRO", exception=e)
            QMessageBox.critical(self, "Erreur", f"Erreur validation DQE PRO: {str(e)}")
            send_telemetry(
                action="validation", mode="PRO",
                sro=self.sro_input.currentText().strip(),
                type=self.type_combo.currentData() if hasattr(self, 'type_combo') else "",
                projet_code="TP" if self.type_combo.currentData() == "T" else "DP",
                verdict="failure", error_msg=str(e),
                elapsed_ms=int((time.monotonic() - validate_start) * 1000),
            )
    
    def _save_all_layers(self, sro, projet_code, user_name):
        """Sauvegarde TOUTES les couches du groupe dans dqejson"""
        return LayerManager.save_layers_to_db(
            self.layer_group, sro, f"DQE_PRO_{sro}",
            projet_code, user_name, _db_manager
        )
