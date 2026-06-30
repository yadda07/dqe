"""
DQE Recover Tab Module

Permet de récupérer et régénérer un DQE à partir des données archivées dans dqejson.
"""

import os
import json
import tempfile
from datetime import datetime
from typing import List, Dict, Any, Optional

from .compat import (
    Qt, QThread, pyqtSignal, QWidget, QVBoxLayout, QHBoxLayout,
    QFormLayout, QGroupBox, QPushButton, QComboBox, QMessageBox,
    QApplication, QTableWidget, QTableWidgetItem, QHeaderView,
    QLabel, QFrame, QLineEdit, QProgressBar, QTextBrowser, QSplitter,
    HEADERVIEW_STRETCH, TABLE_SELECT_ROWS, TABLE_SINGLE_SELECTION,
    QT_VERTICAL, QT_WAIT_CURSOR, MSGBOX_YES, MSGBOX_NO
)
from qgis.core import QgsProject, QgsVectorLayer, QgsFeature, QgsGeometry, Qgis
from qgis.utils import iface

from .layer_manager import LayerManager
from .excel_manager import ExcelManager
from .dqe_utils import _db_manager, _logger


class PreviewWorker(QThread):
    """Worker thread pour charger l'aperçu sans bloquer l'UI"""
    finished = pyqtSignal(list)
    error = pyqtSignal(str)
    
    def __init__(self, sro, projet, nom_dqe, date_validation=None):
        super().__init__()
        self.sro = sro
        self.projet = projet
        self.nom_dqe = nom_dqe
        self.date_validation = date_validation
    
    def run(self):
        try:
            if not _db_manager:
                self.error.emit("DB manager non disponible")
                return
            
            with _db_manager.get_cursor() as cursor:
                if self.date_validation:
                    query = """
                        SELECT categorie, champs
                        FROM dqe.dqejson
                        WHERE sro = %s AND projet = %s AND nom_dqe = %s
                          AND DATE(audit_timestamp) = %s
                        ORDER BY id
                        LIMIT 100
                    """
                    cursor.execute(query, (self.sro, self.projet, self.nom_dqe, self.date_validation))
                else:
                    query = """
                        SELECT categorie, champs
                        FROM dqe.dqejson
                        WHERE sro = %s AND projet = %s AND nom_dqe = %s
                        ORDER BY id
                        LIMIT 100
                    """
                    cursor.execute(query, (self.sro, self.projet, self.nom_dqe))
                
                results = []
                for row in cursor.fetchall():
                    categorie = row[0]
                    # PostgreSQL JSONB retourne directement un objet Python (dict ou list)
                    champs = row[1] if row[1] else {}
                    results.append({'categorie': categorie, 'champs': champs})
                
                self.finished.emit(results)
                
        except Exception as e:
            self.error.emit(str(e))


class DQERecoverTab(QWidget):
    """Interface et logique pour l'onglet DQE Recover"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_dqe = None
        self.dqe_data = []
        self.preview_worker = None
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(4)
        
        # Groupe de sélection
        select_group = QGroupBox("Sélection DQE archivé")
        select_layout = QVBoxLayout(select_group)
        
        # Filtres ligne 1
        filter_layout1 = QHBoxLayout()
        
        # Filtre par type de projet
        self.type_combo = QComboBox()
        self.type_combo.addItem("Tous", None)
        self.type_combo.addItem("PRO Transport", "TP")
        self.type_combo.addItem("PRO Distribution", "DP")
        self.type_combo.addItem("EXE Transport Standard", "TE")
        self.type_combo.addItem("EXE Distribution Standard", "DE")
        self.type_combo.addItem("EXE Transport Travaux", "TT")
        self.type_combo.addItem("EXE Distribution Travaux", "DT")
        self.type_combo.addItem("EXE Transport Blocage", "TB")
        self.type_combo.addItem("EXE Distribution Blocage", "DB")
        self.type_combo.addItem("PGC", "GC")
        filter_layout1.addWidget(QLabel("Type:"))
        filter_layout1.addWidget(self.type_combo)
        
        # Filtre par SRO
        filter_layout1.addWidget(QLabel("SRO:"))
        self.sro_filter = QLineEdit()
        self.sro_filter.setPlaceholderText("Filtrer par SRO...")
        self.sro_filter.setMinimumWidth(150)
        filter_layout1.addWidget(self.sro_filter)
        
        # Bouton rechercher
        self.refresh_btn = QPushButton("Rechercher")
        self.refresh_btn.clicked.connect(self.refresh_dqe_list)
        filter_layout1.addWidget(self.refresh_btn)
        filter_layout1.addStretch()
        
        select_layout.addLayout(filter_layout1)
        
        # Table des DQE disponibles
        self.dqe_table = QTableWidget()
        self.dqe_table.setColumnCount(6)
        self.dqe_table.setHorizontalHeaderLabels([
            "SRO", "Type", "Nom DQE", "Version", "Date création", "Utilisateur"
        ])
        self.dqe_table.horizontalHeader().setSectionResizeMode(HEADERVIEW_STRETCH)
        self.dqe_table.setSelectionBehavior(TABLE_SELECT_ROWS)
        self.dqe_table.setSelectionMode(TABLE_SINGLE_SELECTION)
        self.dqe_table.itemSelectionChanged.connect(self.on_dqe_selected)
        self.dqe_table.setMinimumHeight(200)
        select_layout.addWidget(self.dqe_table)
        
        # Splitter vertical pour redimensionner table/aperçu
        splitter = QSplitter(QT_VERTICAL)
        splitter.addWidget(select_group)
        
        # Groupe aperçu
        preview_group = QGroupBox("Aperçu des données")
        preview_layout = QVBoxLayout(preview_group)
        
        # Barre de chargement (mode indéterminé)
        self.loading_bar = QProgressBar()
        self.loading_bar.setRange(0, 0)
        self.loading_bar.setTextVisible(False)
        self.loading_bar.setMaximumHeight(8)
        self.loading_bar.hide()
        preview_layout.addWidget(self.loading_bar)
        
        # Aperçu HTML avec watermark ARCHIVÉ
        self.preview_browser = QTextBrowser()
        self.preview_browser.setOpenExternalLinks(False)
        self.preview_browser.setMinimumHeight(150)
        self.preview_browser.setStyleSheet("QTextBrowser { background-color: #2b2b2b; color: #e0e0e0; }")
        preview_layout.addWidget(self.preview_browser)
        
        splitter.addWidget(preview_group)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        splitter.setChildrenCollapsible(False)
        splitter.setSizes([200, 400])
        
        layout.addWidget(splitter, 1)
        
        # Groupe actions
        action_group = QGroupBox("Actions")
        action_layout = QHBoxLayout(action_group)
        
        self.recover_excel_btn = QPushButton("Régénérer Excel")
        self.recover_excel_btn.setEnabled(False)
        self.recover_excel_btn.clicked.connect(self.recover_excel)
        action_layout.addWidget(self.recover_excel_btn)
        
        self.recover_layers_btn = QPushButton("Recréer couches QGIS")
        self.recover_layers_btn.setEnabled(False)
        self.recover_layers_btn.clicked.connect(self.recover_layers)
        action_layout.addWidget(self.recover_layers_btn)
        
        self.recover_all_btn = QPushButton("Récupération complète")
        self.recover_all_btn.setEnabled(False)
        self.recover_all_btn.clicked.connect(self.recover_all)
        action_layout.addWidget(self.recover_all_btn)
        
        layout.addWidget(action_group)
        
        # Info
        self.info_label = QLabel("")
        self.info_label.setStyleSheet("color: gray; font-style: italic;")
        layout.addWidget(self.info_label)
        
        # Ne pas charger au démarrage pour éviter le blocage
        self.info_label.setText("Cliquez sur 'Rechercher' pour charger les DQE archivés")
    
    def refresh_dqe_list(self):
        """Actualise la liste des DQE disponibles"""
        try:
            if not _db_manager:
                self.info_label.setText("Connexion base de données non disponible")
                return
            
            self.dqe_table.setRowCount(0)
            self.dqe_data = []
            
            type_filter = self.type_combo.currentData()
            
            sro_filter = self.sro_filter.text().strip()
            
            with _db_manager.get_cursor() as cursor:
                # Construction dynamique de la requête
                conditions = []
                params = []
                
                if type_filter:
                    conditions.append("projet = %s")
                    params.append(type_filter)
                
                if sro_filter:
                    conditions.append("sro ILIKE %s")
                    params.append(f"%{sro_filter}%")
                
                where_clause = ""
                if conditions:
                    where_clause = "WHERE " + " AND ".join(conditions)
                
                # Ne charger QUE les dqe_result (nouveau format simplifié)
                if conditions:
                    where_clause = "WHERE categorie = 'dqe_result' AND " + " AND ".join(conditions)
                else:
                    where_clause = "WHERE categorie = 'dqe_result'"
                
                query = f"""
                    SELECT 
                        sro, projet, nom_dqe,
                        version_projet,
                        audit_timestamp as date_creation,
                        user_name,
                        DATE(audit_timestamp) as date_validation
                    FROM dqe.dqejson
                    {where_clause}
                    ORDER BY audit_timestamp DESC
                    LIMIT 30
                """
                cursor.execute(query, params)
                
                results = cursor.fetchall()
                
                for row in results:
                    self.dqe_data.append({
                        'sro': row[0],
                        'projet': row[1],
                        'nom_dqe': row[2],
                        'version_projet': row[3],
                        'date_creation': row[4],
                        'user_name': row[5],
                        'date_validation': row[6]
                    })
                
                self.dqe_table.setRowCount(len(results))
                for i, row in enumerate(results):
                    self.dqe_table.setItem(i, 0, QTableWidgetItem(str(row[0] or "")))
                    self.dqe_table.setItem(i, 1, QTableWidgetItem(str(row[1] or "")))
                    self.dqe_table.setItem(i, 2, QTableWidgetItem(str(row[2] or "")))
                    self.dqe_table.setItem(i, 3, QTableWidgetItem(str(row[3] or "")))
                    date_str = row[4].strftime("%Y-%m-%d %H:%M") if row[4] else ""
                    self.dqe_table.setItem(i, 4, QTableWidgetItem(date_str))
                    self.dqe_table.setItem(i, 5, QTableWidgetItem(str(row[5] or "")))
                
                self.info_label.setText(f"{len(results)} DQE archivés trouvés")
                
        except Exception as e:
            self.info_label.setText(f"Erreur: {str(e)}")
            print(f"Erreur refresh_dqe_list: {str(e)}")
    
    def on_dqe_selected(self):
        """Appelé quand un DQE est sélectionné"""
        selected_rows = self.dqe_table.selectionModel().selectedRows()
        if not selected_rows:
            self.selected_dqe = None
            self.recover_excel_btn.setEnabled(False)
            self.recover_layers_btn.setEnabled(False)
            self.recover_all_btn.setEnabled(False)
            self.preview_browser.setHtml("")
            return
        
        row_idx = selected_rows[0].row()
        self.selected_dqe = self.dqe_data[row_idx]
        
        # Boutons desactives - fonctionnalite en developpement
        # self.recover_excel_btn.setEnabled(True)
        # self.recover_layers_btn.setEnabled(True)
        # self.recover_all_btn.setEnabled(True)
        
        self.load_preview()
    
    def load_preview(self):
        """Charge l'aperçu des données du DQE sélectionné (asynchrone)"""
        if not self.selected_dqe:
            return
        
        # Arrêter le worker précédent si en cours
        if self.preview_worker and self.preview_worker.isRunning():
            self.preview_worker.quit()
            self.preview_worker.wait(500)
        
        # Afficher animation de chargement
        self.preview_browser.setHtml("<p style='color:#888;'>Chargement...</p>")
        self.loading_bar.show()
        
        # Lancer le worker asynchrone
        self.preview_worker = PreviewWorker(
            self.selected_dqe['sro'],
            self.selected_dqe['projet'],
            self.selected_dqe['nom_dqe'],
            self.selected_dqe.get('date_validation')
        )
        self.preview_worker.finished.connect(self._on_preview_loaded)
        self.preview_worker.error.connect(self._on_preview_error)
        self.preview_worker.start()
    
    def _on_preview_loaded(self, results):
        """Callback quand l'aperçu est chargé - génère rapport HTML"""
        self.loading_bar.hide()
        
        # Générer le rapport HTML avec watermark
        html = self._generate_html_report(results)
        self.preview_browser.setHtml(html)
    
    def _generate_html_report(self, results):
        """Génère un rapport HTML stylisé avec watermark ARCHIVÉ"""
        if not self.selected_dqe:
            return "<p>Aucune donnée</p>"
        
        sro = self.selected_dqe.get('sro', '')
        projet = self.selected_dqe.get('projet', '')
        date_creation = self.selected_dqe.get('date_creation', '')
        date_str = date_creation.strftime('%d/%m/%Y %H:%M') if date_creation else ''
        
        # Extraire les résultats SQL (nouveau et ancien format)
        sql_results = []
        for item in results:
            categorie = item.get('categorie', '')
            champs = item['champs']
            
            # Nouveau format: categorie='dqe_result', champs = tableau JSON
            if categorie == 'dqe_result' and isinstance(champs, list):
                for row in champs:
                    sql_results.append({
                        'designation': row.get('designation', ''),
                        'quantite': row.get('quantite', 0),
                        'unite': row.get('unite', '')
                    })
            # Ancien format: type='sql_result' dans champs
            elif isinstance(champs, dict) and champs.get('type') == 'sql_result':
                sql_results.append({
                    'designation': champs.get('designation', categorie),
                    'quantite': champs.get('quantite', 0),
                    'unite': champs.get('unite', '')
                })
        
        # Générer HTML
        html = f'''
        <html>
        <head>
        <style>
            body {{ 
                font-family: Arial, sans-serif; 
                background-color: #2b2b2b; 
                color: #e0e0e0;
                position: relative;
                margin: 10px;
            }}
            .watermark {{
                position: fixed;
                top: 50%;
                left: 50%;
                transform: translate(-50%, -50%) rotate(-45deg);
                font-size: 48px;
                font-weight: bold;
                color: rgba(255, 0, 0, 0.15);
                pointer-events: none;
                z-index: 1000;
                white-space: nowrap;
            }}
            .header {{
                background: linear-gradient(135deg, #1a5276, #2e86ab);
                padding: 10px 15px;
                border-radius: 5px;
                margin-bottom: 10px;
            }}
            .header h2 {{ margin: 0; color: #fff; font-size: 14px; }}
            .header p {{ margin: 3px 0; color: #bbb; font-size: 11px; }}
            .badge {{
                display: inline-block;
                background: #c0392b;
                color: white;
                padding: 2px 8px;
                border-radius: 3px;
                font-size: 10px;
                margin-left: 10px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 11px;
            }}
            th {{
                background: #34495e;
                color: #fff;
                padding: 6px 8px;
                text-align: left;
                border-bottom: 2px solid #2c3e50;
            }}
            td {{
                padding: 5px 8px;
                border-bottom: 1px solid #444;
            }}
            tr:nth-child(even) {{ background: #333; }}
            tr:hover {{ background: #3d3d3d; }}
            .quantite {{ text-align: right; font-weight: bold; color: #3498db; }}
            .unite {{ color: #888; }}
            .section-header {{
                background: linear-gradient(90deg, #1a5276, #2c3e50);
                font-weight: bold;
                color: #f39c12;
                font-size: 12px;
            }}
            .section-header td {{
                padding: 8px;
                border-bottom: 2px solid #f39c12;
            }}
            .footer {{ 
                margin-top: 10px; 
                padding: 8px; 
                background: #1a1a1a; 
                border-radius: 3px;
                font-size: 10px;
                color: #888;
            }}
        </style>
        </head>
        <body>
        <div class="watermark">ARCHIVÉ</div>
        
        <div class="header">
            <h2>DQE {projet} <span class="badge">ARCHIVÉ</span></h2>
            <p><strong>SRO:</strong> {sro}</p>
            <p><strong>Date:</strong> {date_str}</p>
            <p><strong>Lignes:</strong> {len(sql_results)} résultats SQL</p>
        </div>
        
        <table>
            <tr>
                <th>Désignation</th>
                <th style="width:80px;">Quantité</th>
                <th style="width:50px;">Unité</th>
            </tr>
        '''
        
        for item in sql_results:
            quantite = item['quantite']
            unite = item['unite']
            designation = item['designation']
            
            # Ligne titre si quantité=None ou unite=None
            if quantite is None or unite is None or unite == 'None':
                html += f'''
            <tr class="section-header">
                <td colspan="3">{designation}</td>
            </tr>
                '''
            else:
                if isinstance(quantite, float):
                    quantite = f"{quantite:.1f}" if quantite % 1 else f"{int(quantite)}"
                html += f'''
            <tr>
                <td>{designation}</td>
                <td class="quantite">{quantite}</td>
                <td class="unite">{unite}</td>
            </tr>
                '''
        
        html += '''
        </table>
        
        <div class="footer">
            Document archivé - Régénération possible via les boutons d'action
        </div>
        </body>
        </html>
        '''
        
        return html
    
    def _on_preview_error(self, error_msg):
        """Callback en cas d'erreur de chargement"""
        self.loading_bar.hide()
        self.preview_browser.setHtml(f"<p style='color:#e74c3c;'>Erreur: {error_msg}</p>")
        print(f"Erreur load_preview: {error_msg}")
    
    def get_dqe_data(self) -> List[Dict]:
        """Récupère uniquement la ligne dqe_result du DQE sélectionné"""
        if not self.selected_dqe:
            return []
        
        try:
            with _db_manager.get_cursor() as cursor:
                date_validation = self.selected_dqe.get('date_validation')
                
                # Charger uniquement la ligne dqe_result
                query = """
                    SELECT id, categorie, champs, audit_timestamp
                    FROM dqe.dqejson
                    WHERE sro = %s AND projet = %s AND nom_dqe = %s
                      AND categorie = 'dqe_result'
                      AND DATE(audit_timestamp) = %s
                    LIMIT 1
                """
                cursor.execute(query, (
                    self.selected_dqe['sro'],
                    self.selected_dqe['projet'],
                    self.selected_dqe['nom_dqe'],
                    date_validation
                ))
                
                results = []
                for row in cursor.fetchall():
                    # PostgreSQL JSONB retourne directement un objet Python
                    champs = row[2] if row[2] else []
                    results.append({
                        'id': row[0],
                        'categorie': row[1],
                        'champs': champs,
                        'timestamp': row[3]
                    })
                
                return results
                
        except Exception as e:
            print(f"Erreur get_dqe_data: {str(e)}")
            return []
    
    def _show_dev_in_progress(self):
        """Affiche popup developpement en cours"""
        QMessageBox.information(
            self,
            "Développement en cours",
            "Cette fonctionnalité est en cours de développement.\n\n"
            "Objectif : Reconstruire les données QGIS à la volée\n"
            "à partir de dqe.dqejson."
        )
    
    def recover_excel(self):
        """Régénère le fichier Excel à partir des données archivées"""
        self._show_dev_in_progress()
        return
        
        # Code desactive - en developpement
        if not self.selected_dqe:
            QMessageBox.warning(self, "Récupération", "Veuillez sélectionner un DQE")
            return
        
        try:
            QApplication.setOverrideCursor(QT_WAIT_CURSOR)
            
            dqe_data = self.get_dqe_data()
            if not dqe_data:
                QApplication.restoreOverrideCursor()
                QMessageBox.warning(self, "Récupération", "Aucune donnée trouvée")
                return
            
            # Reconstruire les résultats SQL - LISTE ordonnée (pas dict)
            results = []
            for item in dqe_data:
                categorie = item.get('categorie', '')
                champs = item['champs']
                
                # Nouveau format: categorie='dqe_result', champs = tableau JSON
                if categorie == 'dqe_result' and isinstance(champs, list):
                    for row in champs:
                        results.append({
                            'designation': row.get('designation', ''),
                            'quantite': row.get('quantite', 0),
                            'unite': row.get('unite', '')
                        })
                # Ancien format: type='sql_result' dans champs (compatibilité)
                elif isinstance(champs, dict) and champs.get('type') == 'sql_result':
                    results.append({
                        'designation': champs.get('designation', categorie),
                        'quantite': champs.get('quantite', 0),
                        'unite': champs.get('unite', '')
                    })
            
            print(f"DEBUG Recover: {len(results)} résultats SQL trouvés")
            
            if not results:
                QApplication.restoreOverrideCursor()
                QMessageBox.warning(self, "Récupération", "Aucun résultat SQL trouvé")
                return
            
            # Déterminer le type de rapport
            projet = self.selected_dqe['projet']
            if projet in ('TP', 'DP'):
                report_type = "PRO"
            elif projet in ('TE', 'DE', 'TT', 'DT', 'TB', 'DB'):
                report_type = "EXE"
            else:
                report_type = "PGC"
            
            sro = self.selected_dqe['sro']
            
            # Générer Excel avec toutes les lignes dans l'ordre
            excel_path = self._generate_excel_ordered(results, sro, report_type)
            
            QApplication.restoreOverrideCursor()
            
            if excel_path:
                QMessageBox.information(
                    self, 
                    "Récupération Excel", 
                    f"Fichier Excel régénéré avec succès!\n\n"
                    f"SRO: {sro}\n"
                    f"Type: {report_type}\n"
                    f"Lignes traitées: {len(results)}"
                )
            else:
                QMessageBox.warning(self, "Récupération", "Échec de la génération Excel")
            
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la régénération Excel:\n{str(e)}")
            import traceback
            print(f"Erreur recover_excel: {traceback.format_exc()}")
    
    def _generate_excel_ordered(self, results: list, sro: str, report_type: str) -> Optional[str]:
        """Génère l'Excel avec TOUTES les lignes dans l'ordre exact (comme dqe2)"""
        import shutil
        import time
        import tempfile
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        try:
            temp_dir = tempfile.gettempdir()
            sro_safe = sro.replace('/', '_')
            projet = self.selected_dqe['projet']
            excel_path = os.path.join(temp_dir, f"dqe_RECOVER_{projet}_{sro_safe}_{int(time.time())}.xlsx")
            
            # Créer nouveau workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"DQE {report_type}"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            border = Side(style='thin', color='CCCCCC')
            cell_border = Border(left=border, right=border, top=border, bottom=border)
            
            # En-tête
            sheet.cell(row=1, column=1, value="Désignation").font = header_font
            sheet.cell(row=1, column=1).fill = header_fill
            sheet.cell(row=1, column=2, value="Quantité").font = header_font
            sheet.cell(row=1, column=2).fill = header_fill
            sheet.cell(row=1, column=3, value="Unité").font = header_font
            sheet.cell(row=1, column=3).fill = header_fill
            
            # Largeurs colonnes
            sheet.column_dimensions['A'].width = 60
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 10
            
            # Remplir TOUTES les lignes dans l'ordre
            for i, row_data in enumerate(results, start=2):
                designation = row_data.get('designation', '')
                quantite = row_data.get('quantite', 0)
                unite = row_data.get('unite', '')
                
                sheet.cell(row=i, column=1, value=designation).border = cell_border
                sheet.cell(row=i, column=2, value=quantite).border = cell_border
                sheet.cell(row=i, column=2).alignment = Alignment(horizontal='right')
                sheet.cell(row=i, column=3, value=unite).border = cell_border
            
            # Watermark dans header/footer
            sheet.oddHeader.center.text = f"&\"Arial,Bold\"&14DQE {projet} - {sro} [ARCHIVE]"
            sheet.oddFooter.center.text = f"&\"Arial,Italic\"&8Regenere le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
            print(f"Excel genere: {len(results)} lignes dans l'ordre exact")
            
            workbook.save(excel_path)
            workbook.close()
            
            ExcelManager._open_excel_file(excel_path)
            return excel_path
            
        except Exception as e:
            import traceback
            print(f"Erreur _generate_excel_ordered: {traceback.format_exc()}")
            return None
    
    def recover_layers(self):
        """Recrée les couches QGIS via la fonction SQL dqe_recover"""
        self._show_dev_in_progress()
        return
        
        # Code desactive - en developpement
        if not self.selected_dqe:
            QMessageBox.warning(self, "Récupération", "Veuillez sélectionner un DQE")
            return
        
        try:
            QApplication.setOverrideCursor(QT_WAIT_CURSOR)
            
            sro = self.selected_dqe['sro']
            projet = self.selected_dqe['projet']
            version = self.selected_dqe.get('version_projet', 'dqe')
            sro_safe = sro.replace('/', '_').replace('\\', '_')
            
            # Schéma de destination
            schema_name = f"dqe_recover_{sro_safe}_{datetime.now().strftime('%H%M%S')}"
            schema_name = schema_name[:63].lower()  # Limite PostgreSQL
            
            # Debug: voir combien d'entrées dans dqejson (SANS filtre version)
            with _db_manager.get_cursor() as cursor:
                cursor.execute("""
                    SELECT categorie, jsonb_typeof(champs::jsonb) as json_type,
                           CASE WHEN champs::jsonb->>'type' = 'FeatureCollection' THEN 'FC' ELSE 'autre' END as fc,
                           version_projet
                    FROM dqe.dqejson
                    WHERE sro = %s AND projet = %s
                    ORDER BY audit_timestamp DESC
                """, (sro, projet))
                debug_rows = cursor.fetchall()
                print(f"\n=== DEBUG: {len(debug_rows)} entrees dans dqejson pour {sro}/{projet} ===")
                for r in debug_rows:
                    print(f"  - {r[0]}: {r[1]} / {r[2]} (v={r[3]})")
            
            # Appeler la fonction SQL de reconstruction (SANS filtre version pour tout récupérer)
            tables_created = []
            with _db_manager.get_cursor() as cursor:
                cursor.execute("""
                    SELECT table_name, table_type, geom_type, row_count, message
                    FROM dqe.dqe_recover(%s, %s, %s, %s)
                """, (sro, projet, None, schema_name))
                
                for row in cursor.fetchall():
                    tables_created.append({
                        'name': row[0],
                        'type': row[1],
                        'geom_type': row[2],
                        'count': row[3],
                        'message': row[4]
                    })
                    print(f"  {row[4]}")
            
            if not tables_created:
                QApplication.restoreOverrideCursor()
                QMessageBox.warning(self, "Récupération", "Aucune table créée")
                return
            
            # Créer groupe QGIS
            group_name = f"RECOVER_{projet}_{sro_safe}"
            layer_group = LayerManager.create_layer_group(group_name)
            
            # Charger les tables PostgreSQL comme couches QGIS
            created_count = 0
            from .database_operations import DatabaseOperations
            conn_info = DatabaseOperations.get_db_connection_params()
            
            if not conn_info:
                QApplication.restoreOverrideCursor()
                QMessageBox.warning(self, "Récupération", "Paramètres de connexion non disponibles")
                return
            
            for table_info in tables_created:
                if table_info['count'] == 0:
                    continue
                
                table_name = table_info['name']
                source_type = table_info['type']
                detected_geom = table_info.get('geom_type')
                
                # Construire URI PostgreSQL
                uri = f"dbname='{conn_info['database']}' host='{conn_info['host']}' " \
                      f"port='{conn_info['port']}' user='{conn_info['user']}' " \
                      f"password='{conn_info['password']}' sslmode=disable "
                
                print(f"Table {table_name}: type={source_type}, geom={detected_geom}")
                
                # dqe_result = table sans géométrie
                # FeatureCollection = table avec géométrie détectée depuis WKT
                if detected_geom is None:
                    # Table sans géométrie
                    uri += f"key='id' table=\"{schema_name}\".\"{table_name}\""
                    layer = QgsVectorLayer(uri, table_name, "postgres")
                else:
                    # Table avec géométrie (type détecté: Point, LineString, Polygon, etc.)
                    uri += f"key='id' srid=2154 type={detected_geom} " \
                           f"table=\"{schema_name}\".\"{table_name}\" (geom)"
                    layer = QgsVectorLayer(uri, table_name, "postgres")
                
                if layer and layer.isValid():
                    QgsProject.instance().addMapLayer(layer, False)
                    layer_group.addLayer(layer)
                    created_count += 1
                    print(f"Couche chargée: {table_name} ({layer.featureCount()} features)")
                else:
                    print(f"Couche invalide: {table_name} - URI: {uri}")
            
            QApplication.restoreOverrideCursor()
            
            if iface:
                iface.mapCanvas().refresh()
            
            # Résumé
            summary = "\n".join([f"- {t['name']}: {t['count']} lignes" for t in tables_created if t['count'] > 0])
            QMessageBox.information(
                self,
                "Récupération couches",
                f"Reconstruction terminée!\n\n"
                f"SRO: {sro}\n"
                f"Schéma: {schema_name}\n"
                f"Tables créées: {len(tables_created)}\n"
                f"Couches QGIS: {created_count}\n\n"
                f"{summary}"
            )
            
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la recréation des couches:\n{str(e)}")
            import traceback
            print(f"Erreur recover_layers: {traceback.format_exc()}")
    
    def _create_layer_from_features(self, layer_name: str, features: List[Dict], crs: str = None) -> Optional[QgsVectorLayer]:
        """Crée une couche QGIS à partir de features GeoJSON"""
        try:
            if not features:
                return None
            
            # Déterminer le type de géométrie
            first_geom = features[0].get('geometry', '')
            if 'POINT' in str(first_geom).upper():
                geom_type = 'Point'
            elif 'LINE' in str(first_geom).upper():
                geom_type = 'LineString'
            elif 'POLYGON' in str(first_geom).upper():
                geom_type = 'Polygon'
            else:
                geom_type = 'LineString'  # Par défaut
            
            # Créer couche mémoire
            crs_str = crs or 'EPSG:2154'
            layer = QgsVectorLayer(f"{geom_type}?crs={crs_str}", layer_name, "memory")
            
            if not layer.isValid():
                print(f"Couche mémoire invalide pour {layer_name}")
                return None
            
            provider = layer.dataProvider()
            
            # Ajouter les attributs depuis la premiere feature (helper version-aware)
            first_attrs = features[0].get('attributes', {})
            fields = []
            for attr_name, attr_value in first_attrs.items():
                if isinstance(attr_value, (int, float)):
                    fields.append(LayerManager.create_compatible_field(attr_name, "double"))
                else:
                    fields.append(LayerManager.create_compatible_field(attr_name, "string"))
            provider.addAttributes(fields)
            layer.updateFields()
            
            # Ajouter les features
            qgs_features = []
            for feat_data in features:
                feat = QgsFeature()
                
                # Géométrie
                geom_wkt = feat_data.get('geometry', '')
                if geom_wkt:
                    geom = QgsGeometry.fromWkt(geom_wkt)
                    feat.setGeometry(geom)
                
                # Attributs
                attrs = feat_data.get('attributes', {})
                attr_values = [attrs.get(f.name(), None) for f in layer.fields()]
                feat.setAttributes(attr_values)
                
                qgs_features.append(feat)
            
            provider.addFeatures(qgs_features)
            layer.updateExtents()
            
            return layer
            
        except Exception as e:
            print(f"Erreur _create_layer_from_features: {str(e)}")
            return None
    
    def recover_all(self):
        """Récupération complète: Excel + couches QGIS"""
        self._show_dev_in_progress()
        return
        
        # Code desactive - en developpement
        if not self.selected_dqe:
            QMessageBox.warning(self, "Récupération", "Veuillez sélectionner un DQE")
            return
        
        reply = QMessageBox.question(
            self,
            "Récupération complète",
            f"Régénérer le DQE complet pour:\n\n"
            f"SRO: {self.selected_dqe['sro']}\n"
            f"Type: {self.selected_dqe['projet']}\n\n"
            f"Cela va créer:\n"
            f"- Un fichier Excel\n"
            f"- Les couches QGIS\n\n"
            f"Continuer?",
            MSGBOX_YES | MSGBOX_NO,
            MSGBOX_YES
        )
        
        if reply == MSGBOX_YES:
            self.recover_excel()
            self.recover_layers()
